import os
import json
import traceback
import re
from typing import Dict, List, Any, Optional, Tuple, Union
from datetime import datetime
from collections import defaultdict
import sys
from pathlib import Path
import logging
logger = logging.getLogger(__name__)

# Handle import errors gracefully
try:
    import openpyxl
    from openpyxl.styles import Font, Fill, Border, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    logger.warning(f"Warning: openpyxl import failed - Excel metadata extraction will not be available: {str(e)}")

try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    logger.warning("Warning: xlwings not available - display values will not be extracted")


# Add the project parent to Python path
project_root = Path(__file__).parent.parent.absolute()
sys.path.append(str(project_root))
from storage.excel_metadata_storage import ExcelMetadataStorage


class ExcelDependencyExtractor:
    """Helper class for extracting cell dependencies from Excel formulas"""
    
    def __init__(self):
        self.dependency_map = defaultdict(set)  # cell -> cells it depends on
        self.dependent_map = defaultdict(set)   # cell -> cells that depend on it
    
    def column_string_to_number(self, col_str):
        """Convert column string (A, B, ..., AA, AB, etc.) to number"""
        num = 0
        for char in col_str:
            num = num * 26 + (ord(char.upper()) - ord('A') + 1)
        return num
    
    def number_to_column_string(self, col_num):
        """Convert column number to string (1->A, 2->B, ..., 27->AA, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def parse_cell_address(self, cell_ref):
        """Parse cell address like 'A1', '$A$1', 'A$1', '$A1' into (column, row)"""
        clean_ref = cell_ref.replace('$', '')
        match = re.match(r'^([A-Z]+)(\d+)$', clean_ref.upper())
        if match:
            col_str, row_str = match.groups()
            col_num = self.column_string_to_number(col_str)
            row_num = int(row_str)
            return col_num, row_num
        return None, None
    
    def extract_ranges_and_expand(self, formula, current_sheet):
        """Extract cell ranges (with colons) and expand them to individual cells"""
        if not formula or not formula.startswith('='):
            return set(), formula
        
        expanded_cells = set()
        modified_formula = formula
        
        # Pattern to match ranges with optional sheet references
        range_pattern = r"(?:(?:'([^']+)'|([^'!\s(),\+\-\*/]+))!)?(\$?[A-Z]+\$?\d+):(\$?[A-Z]+\$?\d+)"
        
        range_matches = re.finditer(range_pattern, formula.upper())
        
        for match in range_matches:
            quoted_sheet, unquoted_sheet, start_cell, end_cell = match.groups()
            sheet_name = quoted_sheet or unquoted_sheet or current_sheet
            
            start_col, start_row = self.parse_cell_address(start_cell)
            end_col, end_row = self.parse_cell_address(end_cell)
            
            if None in (start_col, start_row, end_col, end_row):
                continue
            
            # Ensure start <= end
            if start_col > end_col:
                start_col, end_col = end_col, start_col
            if start_row > end_row:
                start_row, end_row = end_row, start_row
            
            # Limit range expansion to prevent memory issues
            max_cells = 10000  # Adjust as needed
            range_size = (end_row - start_row + 1) * (end_col - start_col + 1)
            
            if range_size > max_cells:
                logger.warning(f"Warning: Range {start_cell}:{end_cell} too large ({range_size} cells), skipping expansion")
                continue
            
            # Expand the range
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    col_str = self.number_to_column_string(col)
                    cell_addr = f"{sheet_name}!{col_str}{row}"
                    expanded_cells.add(cell_addr)
            
            # Remove this range from the formula
            full_range = match.group(0)
            modified_formula = modified_formula.replace(full_range, f"__RANGE_PLACEHOLDER_{len(expanded_cells)}__", 1)
        
        return expanded_cells, modified_formula
    
    def extract_individual_cells(self, formula, current_sheet):
        """Extract individual cell references (not ranges) from formula"""
        if not formula or not formula.startswith('='):
            return set()
        
        individual_cells = set()
        cell_pattern = r"(?:(?:'([^']+)'|([^'!\s(),\+\-\*/]+))!)?(\$?[A-Z]+\$?\d+)(?!:)"
        
        cell_matches = re.finditer(cell_pattern, formula.upper())
        
        for match in cell_matches:
            quoted_sheet, unquoted_sheet, cell_ref = match.groups()
            sheet_name = quoted_sheet or unquoted_sheet or current_sheet
            
            col, row = self.parse_cell_address(cell_ref)
            if col is not None and row is not None:
                cell_addr = f"{sheet_name}!{self.number_to_column_string(col)}{row}"
                individual_cells.add(cell_addr)
        
        return individual_cells
    
    def extract_all_cell_references(self, formula, current_sheet):
        """Extract all cell references from formula: ranges + individual cells"""
        if not formula or not formula.startswith('='):
            return set()
        
        all_references = set()
        
        try:
            # Extract and expand ranges
            range_cells, modified_formula = self.extract_ranges_and_expand(formula, current_sheet)
            all_references.update(range_cells)
            
            # Extract individual cells
            individual_cells = self.extract_individual_cells(modified_formula, current_sheet)
            all_references.update(individual_cells)
        except Exception as e:
            logger.warning(f"Warning: Error extracting references from formula '{formula}': {str(e)}")
        
        return all_references
    
    def build_dependency_maps(self, all_cells_metadata):
        """Build complete dependency maps from all cell metadata"""
        logger.info("Building dependency maps...")
        
        # Reset maps
        self.dependency_map.clear()
        self.dependent_map.clear()
        
        formula_count = 0
        total_dependencies = 0
        
        for cell_address, cell_data in all_cells_metadata.items():
            formula = cell_data.get('formula')
            if not formula:
                continue
            
            formula_count += 1
            sheet_name = cell_data.get('sheet', cell_address.split('!')[0] if '!' in cell_address else 'Sheet1')
            
            # Extract all cell references
            try:
                precedents = self.extract_all_cell_references(formula, sheet_name)
                
                # Build bidirectional mapping
                for precedent in precedents:
                    self.dependency_map[cell_address].add(precedent)
                    self.dependent_map[precedent].add(cell_address)
                
                total_dependencies += len(precedents)
                
            except Exception as e:
                logger.warning(f"Warning: Error processing dependencies for {cell_address}: {str(e)}")
                continue
        
        logger.info(f"Built dependency maps for {formula_count} formula cells with {total_dependencies} total dependencies")
        return self.dependency_map, self.dependent_map

class ExcelMetadataExtractor:
    
    def __init__(self, 
        workbook_path: Optional[str] = None, 
        use_storage: Optional[bool] = True,
        storage: Optional[ExcelMetadataStorage] = None):
        """
        Initialize the metadata extractor with enhanced error handling and dependency extraction.
        
        Args:
            workbook_path: Path to the Excel file (optional, can be set later)
            use_storage: Whether to use storage for metadata extraction
            storage: Optional storage instance to use
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl is required but not installed. Please install with: pip install openpyxl")
            
        self.workbook_path = None
        self.workbook = None
        self.workbook_values = None
        self.workbook_for_charts = None
        self.xlwings_extractor = None
        self.dependency_extractor = ExcelDependencyExtractor()
        self._initialize_xlwings()
        logger.info("ExcelMetadataExtractor initialized, initialized storage for extractor")
        self.use_storage = use_storage
        if self.use_storage:
            try:
                self.storage = storage or ExcelMetadataStorage()
                logger.info("Storage initialized, ExcelMetadataExtractor class initialization complete.")
            except Exception as e:
                logger.error(f"Error initializing storage: {str(e)}")
                self.storage = None
                self.use_storage = False
            

    def _initialize_xlwings(self) -> None:
        """Initialize xlwings extractor if available."""
        if not XLWINGS_AVAILABLE:
            logger.warning("Warning: xlwings not available - display values will not be extracted")
            return
            
        try:
            from .xlwings_extractor import XlwingsMetadataExtractor
            self.xlwings_extractor = XlwingsMetadataExtractor()
        except ImportError as e:
            logger.warning(f"Warning: Failed to initialize xlwings extractor: {str(e)}")
        except Exception as e:
            logger.warning(f"Warning: Unexpected error initializing xlwings: {str(e)}")
            logger.error(traceback.format_exc())

    def open_workbook(self, workbook_path: Optional[str] = None) -> None:
        """Open the Excel workbook with comprehensive error handling."""
        try:
            if workbook_path:
                self.workbook_path = os.path.abspath(workbook_path)
                
            if not self.workbook_path:
                logger.error("No workbook path specified")
                
            if not os.path.exists(self.workbook_path):
                logger.error(f"Workbook not found: {self.workbook_path}")
                
            if not os.path.isfile(self.workbook_path):
                logger.error(f"Path is not a file: {self.workbook_path}")
                
            # Open workbooks with explicit error handling
            try:
                self.workbook = openpyxl.load_workbook(
                    self.workbook_path, 
                    data_only=False,
                    read_only=True,
                    keep_links=False
                )
            except Exception as e:
                logger.error(f"Failed to open workbook for formulas: {str(e)}")
                
            try:
                self.workbook_values = openpyxl.load_workbook(
                    self.workbook_path,
                    data_only=True,
                    read_only=True,
                    keep_links=False
                )
            except Exception as e:
                if self.workbook:
                    self.workbook.close()
                    self.workbook = None
                logger.error(f"Failed to open workbook for values: {str(e)}")
            
            # Open workbook for charts (read_only=False to access drawing objects)
            try:
                self.workbook_for_charts = openpyxl.load_workbook(
                    self.workbook_path,
                    data_only=False,
                    read_only=False,  # Must be False to access charts
                    keep_links=False
                )
            except Exception as e:
                if self.workbook:
                    self.workbook.close()
                    self.workbook = None
                if self.workbook_values:
                    self.workbook_values.close()
                    self.workbook_values = None
                logger.error(f"Failed to open workbook for charts: {str(e)}")
                
        except Exception as e:
            self.close()
            logger.error(f"Error opening workbook: {str(e)}")

    def close(self) -> None:
        """Safely close all open workbooks and clean up resources."""
        try:
            if self.workbook:
                self.workbook.close()
        except Exception as e:
            logger.warning(f"Warning: Error closing workbook: {str(e)}")
        finally:
            self.workbook = None
            
        try:
            if self.workbook_values:
                self.workbook_values.close()
        except Exception as e:
            logger.warning(f"Warning: Error closing values workbook: {str(e)}")
        finally:
            self.workbook_values = None
            
        try:
            if self.workbook_for_charts:
                self.workbook_for_charts.close()
        except Exception as e:
            logger.warning(f"Warning: Error closing charts workbook: {str(e)}")
        finally:
            self.workbook_for_charts = None

    def __enter__(self):
        """Context manager entry."""
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - ensures resources are cleaned up."""
        self.close()
        return False  # Don't suppress exceptions

    def extract_workbook_metadata(
        self,
        workbook_path: Optional[str] = None,
        output_path: Optional[str] = None,
        max_rows_per_sheet: int = 100,
        max_cols_per_sheet: int = 50,
        include_display_values: bool = False,
        include_dependencies: bool = True,
        force_extract: bool = False
    ) -> Tuple[Dict[str, Any], Dict[str, str]]:
        """
        Extract metadata with comprehensive error handling and dependency analysis.
        Optionally checks storage first to avoid re-extracting unchanged workbooks.
        
        Args:
            workbook_path: Path to the Excel file
            output_path: Optional path to save the metadata
            max_rows_per_sheet: Maximum rows to extract per sheet
            max_cols_per_sheet: Maximum columns to extract per sheet
            include_display_values: Whether to include display values using xlwings
            include_dependencies: Whether to include cell dependency analysis
            force_extract: If True, always extract even if unchanged
            
        Returns:
            Tuple of (metadata_dict, display_values_dict)
        """
        try:

            # Ensure we have a storage instance if needed
            if self.use_storage:
                if not hasattr(self, 'storage'):
                    self.storage = ExcelMetadataStorage()
                
                # Normalize the file path
                normalized_path = str(Path(workbook_path).resolve()) if workbook_path else None
                
                # Check storage first if not forcing extraction
                if not force_extract and normalized_path and hasattr(self, 'storage'):
                    try:
                        # Get the latest version from storage
                        latest_version = self.storage.get_latest_version(normalized_path)
                        if latest_version:
                            # Get the file's current hash
                            current_hash = self._calculate_file_hash(normalized_path)
                            
                            # If the file hasn't changed, return stored metadata
                            if latest_version.get('file_hash') == current_hash:
                                logger.info(f"Using cached metadata for {normalized_path}")
                                stored_metadata = latest_version.get('full_metadata_json')
                                if stored_metadata:
                                    try:
                                        metadata = json.loads(stored_metadata)
                                        # Validate the metadata structure
                                        if self._is_valid_metadata(metadata):
                                            logger.info(f"Using cached metadata for {normalized_path}")
                                            return metadata
                                        else:
                                            logger.info("Cached metadata is invalid, forcing re-extraction")
                                    except (json.JSONDecodeError, TypeError) as e:
                                        logger.warning(f"Error parsing stored metadata: {str(e)}")
                                else:
                                    logger.info("No metadata found in storage, forcing extraction")
                    except Exception as e:
                        logger.warning(f"Warning: Error checking storage for existing metadata: {str(e)}, forcing extraction")
            
            # If we get here, either:
            # 1. No existing version found in storage
            # 2. File has changed
            # 3. force_extract is True
            # 4. There was an error checking storage
            # 5. use_storage is set to false and the class is oconfigured not to use storage

            # Get metadata using openpyxl
            metadata = self.extract_workbook_metadata_openpyxl(
                workbook_path,
                max_rows_per_sheet,
                max_cols_per_sheet,
                include_dependencies
            )
            
            # Get display values if requested and xlwings is available
            display_values = {}
            if include_display_values and self.xlwings_extractor:
                try:
                    display_values = self.xlwings_extractor._extract_display_values_xlwings(
                        workbook_path or self.workbook_path,
                        metadata
                    )
                except Exception as e:
                    logger.warning(f"Warning: Failed to extract display values: {str(e)}")
                    display_values = {"error": str(e)}


            if self.use_storage:
                # Save to storage
                if normalized_path and hasattr(self, 'storage'):
                    try:
                        # Create or update workbook in storage
                        workbook_id = self.storage.create_or_update_workbook(normalized_path)
                        
                        # Create a new version with the extracted metadata
                        version_id = self.storage.create_new_version(
                            file_path=normalized_path,
                            change_description="Initial extraction" if not latest_version else "Updated extraction",
                            full_metadata_json=json.dumps(metadata),
                            store_file_blob=True                            
                        )
                        logger.info(f"Stored metadata in storage for version {version_id}")
                    except Exception as e:
                        logger.warning(f"Warning: Failed to store metadata in storage: {str(e)}")
                        import traceback
                        traceback.print_exc()

            
            return metadata, display_values
            
        except Exception as e:
            error_msg = f"Failed to extract workbook metadata: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            

    def extract_workbook_metadata_openpyxl(
        self,
        workbook_path: Optional[str] = None,
        max_rows_per_sheet: int = 100,
        max_cols_per_sheet: int = 50,
        include_dependencies: bool = True
    ) -> Dict[str, Any]:
        """
        Extract metadata using openpyxl with comprehensive error handling and dependency analysis.
        """
        try:
            if workbook_path or not self.workbook:
                self.open_workbook(workbook_path)
                
            if not self.workbook:
                logger.error("Workbook is not open")
                
            # Get basic workbook info
            workbook_metadata = {
                "extractedAt": datetime.now().isoformat(),
                "workbookPath": self.workbook_path,
                "workbookName": os.path.basename(self.workbook_path) if self.workbook_path else "Unknown",
                "activeSheet": self.workbook.active.title if self.workbook.active else None,
                "totalSheets": len(self.workbook.worksheets),
                "sheetNames": [sheet.title for sheet in self.workbook.worksheets],
                "themeColors": self._get_theme_colors(),
                "sheets": [],
                "includeDependencies": include_dependencies
            }
            
            # First pass: Extract all cell metadata (without dependencies)
            all_cells_metadata = {}
            
            # Extract metadata for each sheet
            for sheet in self.workbook.worksheets:
                try:
                    sheet_metadata, sheet_cells = self._extract_sheet_metadata_with_cells(
                        sheet,
                        max_rows_per_sheet,
                        max_cols_per_sheet
                    )
                    workbook_metadata["sheets"].append(sheet_metadata)
                    
                    # Collect all cells for dependency analysis
                    all_cells_metadata.update(sheet_cells)
                    
                except Exception as e:
                    error_msg = f"Error processing sheet '{sheet.title}': {str(e)}"
                    logger.warning(f"Warning: {error_msg}")
                    logger.error(traceback.format_exc())
                    workbook_metadata["sheets"].append({
                        "name": sheet.title,
                        "error": error_msg,
                        "isEmpty": True
                    })
            
            # Second pass: Build dependency relationships if requested
            if include_dependencies and all_cells_metadata:
                logger.info("Building dependency relationships...")
                try:
                    dependency_map, dependent_map = self.dependency_extractor.build_dependency_maps(all_cells_metadata)
                    
                    # Third pass: Update all cells with dependency information
                    self._update_cells_with_dependencies(
                        workbook_metadata,
                        all_cells_metadata,
                        dependency_map,
                        dependent_map
                    )
                    
                    # Add dependency summary
                    self._add_dependency_summary(workbook_metadata, dependency_map, dependent_map, all_cells_metadata)
                    
                except Exception as e:
                    error_msg = f"Error building dependencies: {str(e)}"
                    logger.warning(f"Warning: {error_msg}")
                    logger.error(traceback.format_exc())
                    workbook_metadata["dependencyError"] = error_msg
            
            return workbook_metadata
            
        except Exception as e:
            error_msg = f"Failed to extract workbook metadata: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())

    def _extract_sheet_metadata_with_cells(
        self,
        sheet,
        max_rows: int,
        max_cols: int
    ) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        """
        Extract metadata from a single sheet and return both sheet metadata and cell dictionary.
        """
        sheet_metadata = {
            "name": sheet.title,
            "error": None,
            "isEmpty": True,
            "rowCount": 0,
            "columnCount": 0,
            "cellData": [],
            "tables": [],
            "namedRanges": []
        }
        
        sheet_cells = {}  # For dependency analysis: {full_address: cell_metadata}
        
        try:
            # Get actual dimensions
            actual_row_count = min(sheet.max_row or 0, 1048576)  # Excel max rows
            actual_col_count = min(sheet.max_column or 0, 16384)  # Excel max columns
            
            # Check if sheet is empty
            if actual_row_count == 0 or actual_col_count == 0:
                sheet_metadata["isEmpty"] = True
                return sheet_metadata, sheet_cells
                
            sheet_metadata.update({
                "isEmpty": False,
                "rowCount": actual_row_count,
                "columnCount": actual_col_count,
                "extractedRowCount": min(actual_row_count, max_rows),
                "extractedColumnCount": min(actual_col_count, max_cols)
            })
            
            # Extract cell data
            try:
                sheet_metadata["cellData"], sheet_cells = self._extract_cell_data_with_addresses(
                    sheet,
                    sheet_metadata["extractedRowCount"],
                    sheet_metadata["extractedColumnCount"]
                )
            except Exception as e:
                sheet_metadata["error"] = f"Error extracting cell data: {str(e)}"
                logger.warning(f"Warning: {sheet_metadata['error']}")
                logger.error(traceback.format_exc())
                sheet_metadata["cellData"] = []
            
            # Extract tables and named ranges
            try:
                sheet_metadata["tables"] = self._extract_tables_metadata(sheet)
            except Exception as e:
                sheet_metadata["error"] = sheet_metadata.get("error", "") + f" Tables: {str(e)}"
                logger.warning(f"Warning: Error extracting tables: {str(e)}")
                logger.error(traceback.format_exc())
                sheet_metadata["tables"] = []
            
            # Extract charts
            try:
                sheet_metadata["charts"] = self._extract_charts_from_sheet(sheet)
            except Exception as e:
                sheet_metadata["error"] = sheet_metadata.get("error", "") + f" Charts: {str(e)}"
                logger.warning(f"Warning: Error extracting charts: {str(e)}")
                logger.error(traceback.format_exc())
                sheet_metadata["charts"] = []

            # TODO: Extract named ranges
            
            return sheet_metadata, sheet_cells
            
        except Exception as e:
            error_msg = f"Critical error extracting sheet '{sheet.title}': {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            sheet_metadata.update({
                "error": error_msg,
                "isEmpty": True
            })
            return sheet_metadata, sheet_cells

    def _extract_cell_data_with_addresses(
        self,
        sheet,
        row_count: int,
        col_count: int
    ) -> Tuple[List[List[Dict[str, Any]]], Dict[str, Any]]:
        """
        Extract cell data and return both 2D array and address-keyed dictionary.
        """
        cell_data = []
        cell_dict = {}  # For dependency analysis
        
        try:
            for row_idx in range(1, row_count + 1):
                row_data = []
                
                for col_idx in range(1, col_count + 1):
                    try:
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell_metadata = self._extract_complete_cell_metadata(cell, row_idx, col_idx, sheet.title)
                        row_data.append(cell_metadata)
                        
                        # Add to dictionary for dependency analysis if has content
                        if cell_metadata.get('value') is not None or cell_metadata.get('formula'):
                            full_address = f"{sheet.title}!{cell_metadata['address']}"
                            cell_dict[full_address] = cell_metadata
                            
                    except Exception as e:
                        logger.warning(f"Warning: Error processing cell ({row_idx},{col_idx}): {str(e)}")
                        logger.error(traceback.format_exc())
                        row_data.append({
                            "row": row_idx,
                            "column": col_idx,
                            "address": f"{get_column_letter(col_idx)}{row_idx}",
                            "error": str(e)
                        })
                
                cell_data.append(row_data)
                
        except Exception as e:
            logger.error(f"Error extracting cell data: {str(e)}")
            logger.error(traceback.format_exc())
            
            
        return cell_data, cell_dict

    def _extract_complete_cell_metadata(
        self,
        cell,
        row: int,
        col: int,
        sheet_name: str
    ) -> Dict[str, Any]:
        """
        Extract complete cell metadata including dependency placeholders.
        """
        try:
            # Handle EmptyCell objects
            if hasattr(cell, '__class__') and cell.__class__.__name__ == 'EmptyCell':
                return {
                    "row": row,
                    "column": col,
                    "address": f"{get_column_letter(col)}{row}",
                    "sheet": sheet_name,
                    "value": None,
                    "formula": None,
                    "formatting": {},
                    # Dependency information (will be filled later)
                    "directPrecedents": [],
                    "directDependents": [],
                    "precedentCount": 0,
                    "dependentCount": 0,
                    "totalConnections": 0
                }
                
            # Get corresponding cell from values workbook
            value_sheet = self.workbook_values[cell.parent.title]
            value_cell = value_sheet.cell(row=row, column=col)
            
            cell_metadata = {
                "row": row,
                "column": col,
                "address": f"{get_column_letter(col)}{row}",
                "sheet": sheet_name,
                "value": self._serialize_value(value_cell.value),
                "formula": self._get_cell_formula(cell),
                "formatting": self._extract_complete_cell_formatting(cell),
                # Dependency information (will be filled later)
                "directPrecedents": [],
                "directDependents": [],
                "precedentCount": 0,
                "dependentCount": 0,
                "totalConnections": 0
            }
            
            return cell_metadata
            
        except Exception as e:
            logger.error(f"Error extracting metadata for cell ({row},{col}): {str(e)}")
            logger.error(traceback.format_exc())
            return {
                "row": row,
                "column": col,
                "address": f"{get_column_letter(col)}{row}",
                "sheet": sheet_name,
                "error": str(e),
                # Dependency information
                "directPrecedents": [],
                "directDependents": [],
                "precedentCount": 0,
                "dependentCount": 0,
                "totalConnections": 0
            }

    def _update_cells_with_dependencies(
        self,
        workbook_metadata: Dict[str, Any],
        all_cells_metadata: Dict[str, Any],
        dependency_map: Dict[str, set],
        dependent_map: Dict[str, set]
    ):
        """Update all cells in workbook metadata with dependency information."""
        logger.info("Updating cells with dependency information...")
        
        for sheet_data in workbook_metadata["sheets"]:
            if sheet_data.get("isEmpty") or "cellData" not in sheet_data:
                continue
                
            sheet_name = sheet_data["name"]
            
            for row_data in sheet_data["cellData"]:
                for cell_metadata in row_data:
                    if "address" in cell_metadata:
                        full_address = f"{sheet_name}!{cell_metadata['address']}"
                        
                        # Get dependency information
                        precedents = dependency_map.get(full_address, set())
                        dependents = dependent_map.get(full_address, set())
                        
                        # Update cell metadata
                        cell_metadata.update({
                            "directPrecedents": list(precedents),
                            "directDependents": list(dependents),
                            "precedentCount": len(precedents),
                            "dependentCount": len(dependents),
                            "totalConnections": len(precedents) + len(dependents)
                        })

    def _add_dependency_summary(
        self,
        workbook_metadata: Dict[str, Any],
        dependency_map: Dict[str, set],
        dependent_map: Dict[str, set],
        all_cells_metadata: Dict[str, Any]
    ):
        """Add dependency summary to workbook metadata."""
        
        total_cells = len(all_cells_metadata)
        formula_cells = sum(1 for data in all_cells_metadata.values() if data.get('formula'))
        total_dependencies = sum(len(deps) for deps in dependency_map.values())
        
        # Find most connected cells
        cell_connections = {}
        for cell_addr in all_cells_metadata:
            precedent_count = len(dependency_map.get(cell_addr, set()))
            dependent_count = len(dependent_map.get(cell_addr, set()))
            cell_connections[cell_addr] = precedent_count + dependent_count
        
        most_connected = sorted(cell_connections.items(), key=lambda x: x[1], reverse=True)[:10]
        most_precedents = sorted(
            [(cell, len(deps)) for cell, deps in dependency_map.items()],
            key=lambda x: x[1], reverse=True
        )[:10]
        most_dependents = sorted(
            [(cell, len(deps)) for cell, deps in dependent_map.items()],
            key=lambda x: x[1], reverse=True
        )[:10]
        
        # Add summary
        workbook_metadata["dependencySummary"] = {
            "totalCells": total_cells,
            "formulaCells": formula_cells,
            "valueCells": total_cells - formula_cells,
            "totalDependencies": total_dependencies,
            "avgDependenciesPerCell": total_dependencies / total_cells if total_cells > 0 else 0,
            "mostConnectedCells": [{"cell": cell, "connections": count} for cell, count in most_connected if count > 0],
            "mostComplexFormulas": [{"cell": cell, "precedents": count} for cell, count in most_precedents if count > 0],
            "mostReferencedCells": [{"cell": cell, "dependents": count} for cell, count in most_dependents if count > 0]
        }
        
        logger.info(f"Added dependency summary: {total_dependencies} total dependencies across {formula_cells} formula cells")

    def _get_cell_formula(self, cell) -> Optional[str]:
        """Safely get cell formula."""
        try:
            if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                return cell.value
            return None
        except Exception:
            return None

    def _extract_complete_cell_formatting(self, cell) -> Dict[str, Any]:
        """
        Extract all cell formatting with comprehensive error handling.
        """
        formatting = {}
        
        try:
            # Font properties
            try:
                if cell.font:
                    formatting["font"] = {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "color": self._get_color_hex(cell.font.color),
                        "underline": cell.font.underline,
                        "strikethrough": cell.font.strikethrough,
                        "vertAlign": cell.font.vertAlign
                    }
            except Exception as e:
                formatting["font"] = {"error": str(e)}
                
            # Fill properties
            try:
                if cell.fill:
                    formatting["fill"] = {
                        "fillType": cell.fill.fill_type,
                        "startColor": self._get_color_hex(cell.fill.start_color),
                        "endColor": self._get_color_hex(cell.fill.end_color),
                        "patternType": cell.fill.patternType
                    }
            except Exception as e:
                formatting["fill"] = {"error": str(e)}
                
            # Number format
            try:
                formatting["numberFormat"] = cell.number_format
            except Exception as e:
                formatting["numberFormat"] = {"error": str(e)}
                
            # Alignment
            try:
                if cell.alignment:
                    alignment_data = {
                        "horizontal": getattr(cell.alignment, 'horizontal', None),
                        "vertical": getattr(cell.alignment, 'vertical', None),
                        "wrapText": getattr(cell.alignment, 'wrap_text', None),
                        "shrinkToFit": getattr(cell.alignment, 'shrink_to_fit', None),
                        "indent": getattr(cell.alignment, 'indent', None),
                        "textRotation": getattr(cell.alignment, 'text_rotation', None)
                    }
                    # Only include properties that exist and don't cause errors
                    if hasattr(cell.alignment, 'justify_last_line'):
                        alignment_data["justifyLastLine"] = cell.alignment.justify_last_line
                    if hasattr(cell.alignment, 'relative_indent'):
                        alignment_data["relativeIndent"] = cell.alignment.relative_indent
                    
                    formatting["alignment"] = alignment_data
            except Exception as e:
                # Skip alignment if it causes errors
                pass
                
            # Border properties
            try:
                if cell.border:
                    border_data = {
                        "left": self._get_border_info(getattr(cell.border, 'left', None)),
                        "right": self._get_border_info(getattr(cell.border, 'right', None)),
                        "top": self._get_border_info(getattr(cell.border, 'top', None)),
                        "bottom": self._get_border_info(getattr(cell.border, 'bottom', None))
                    }
                    
                    # Add optional properties if they exist
                    if hasattr(cell.border, 'diagonal'):
                        border_data["diagonal"] = self._get_border_info(cell.border.diagonal)
                    if hasattr(cell.border, 'diagonal_up'):
                        border_data["diagonalUp"] = cell.border.diagonal_up
                    if hasattr(cell.border, 'diagonal_down'):
                        border_data["diagonalDown"] = cell.border.diagonal_down
                    if hasattr(cell.border, 'outline'):
                        border_data["outline"] = cell.border.outline
                    if hasattr(cell.border, 'start'):
                        border_data["start"] = self._get_border_info(cell.border.start)
                    if hasattr(cell.border, 'end'):
                        border_data["end"] = self._get_border_info(cell.border.end)
                    
                    formatting["borders"] = border_data
            except Exception as e:
                # Skip borders if they cause errors
                pass
                
            # Protection
            try:
                if cell.protection:
                    formatting["protection"] = {
                        "locked": cell.protection.locked,
                        "hidden": cell.protection.hidden
                    }
            except Exception as e:
                formatting["protection"] = {"error": str(e)}
                
            # Comments - only try if cell has comment attribute
            try:
                if hasattr(cell, 'comment') and cell.comment:
                    formatting["comment"] = {
                        "text": getattr(cell.comment, 'text', None),
                        "author": getattr(cell.comment, 'author', None),
                        "width": getattr(cell.comment, 'width', None),
                        "height": getattr(cell.comment, 'height', None)
                    }
            except Exception as e:
                # Skip comments if they cause errors
                pass
                
            # Hyperlink - only try if cell has hyperlink attribute
            try:
                if hasattr(cell, 'hyperlink') and cell.hyperlink:
                    formatting["hyperlink"] = {
                        "target": getattr(cell.hyperlink, 'target', None),
                        "tooltip": getattr(cell.hyperlink, 'tooltip', None),
                        "display": getattr(cell.hyperlink, 'display', None)
                    }
            except Exception as e:
                # Skip hyperlinks if they cause errors
                pass
                
            # Data type
            try:
                formatting["dataType"] = cell.data_type
            except Exception as e:
                formatting["dataType"] = {"error": str(e)}
                
            # Merged cells - only try if worksheet has merged_cells attribute
            try:
                if hasattr(cell.parent, 'merged_cells') and cell.parent.merged_cells:
                    merged_ranges = cell.parent.merged_cells.ranges
                    is_merged = False
                    merge_range = None
                    
                    for merged_range in merged_ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            merge_range = str(merged_range)
                            break
                            
                    formatting["merged"] = {
                        "isMerged": is_merged,
                        "mergeRange": merge_range
                    }
            except Exception as e:
                # Skip merged cells if they cause errors
                pass
                
        except Exception as e:
            formatting["error"] = f"Error extracting formatting: {str(e)}"
            logger.warning(f"Warning: {formatting['error']}")
            logger.error(traceback.format_exc())
            
        return formatting

    def _get_color_hex(self, color) -> Optional[str]:
        """Safely convert color to hex string."""
        if color is None:
            return None
        
        try:
            # Handle RGB object (newer openpyxl)
            if hasattr(color, 'rgb'):
                if hasattr(color, 'rgb') and color.rgb is not None:
                    # Handle RGB object directly
                    if hasattr(color.rgb, 'rgb'):
                        return f"#{color.rgb.rgb[2:]}"  # Skip alpha channel
                    # Handle string RGB
                    elif isinstance(color.rgb, str):
                        if color.rgb.startswith('FF'):
                            return f"#{color.rgb[2:]}"
                        return f"#{color.rgb}"
                return None
                
            # Handle indexed color
            if hasattr(color, 'indexed') and color.indexed is not None:
                return f"indexed_{color.indexed}"
                
            # Handle theme color
            if hasattr(color, 'theme') and color.theme is not None:
                return f"theme_{color.theme}"
                
            # Handle auto color
            if hasattr(color, 'auto') and color.auto:
                return "auto"
                
            return str(color)
        except Exception as e:
            logger.warning(f"Warning: Error converting color to hex: {str(e)}")
            logger.error(traceback.format_exc())
            return None

    def _get_border_info(self, border_side) -> Dict[str, Any]:
        """Safely get border information."""
        try:
            if not border_side:
                return {"style": None, "color": None}
                
            return {
                "style": border_side.style,
                "color": self._get_color_hex(border_side.color)
            }
        except Exception as e:
            logger.warning(f"Warning: Error getting border info: {str(e)}")
            logger.error(traceback.format_exc())
            return {"style": None, "color": None, "error": str(e)}

    def _serialize_value(self, value: Any) -> Any:
        """Safely serialize a cell value."""
        try:
            if value is None:
                return None
            elif isinstance(value, (str, int, float, bool)):
                return value
            elif hasattr(value, 'isoformat'):  # datetime objects
                return value.isoformat()
            else:
                return str(value)
        except Exception as e:
            logger.warning(f"Warning: Error serializing value: {str(e)}")
            logger.error(traceback.format_exc())
            return None

    def _extract_tables_metadata(self, sheet) -> List[Dict[str, Any]]:
        """Safely extract table metadata."""
        tables = []
        
        try:
            if not hasattr(sheet, 'tables'):
                return tables
                
            for table_name, table in sheet.tables.items():
                try:
                    table_data = {
                        "name": table_name,
                        "displayName": table.displayName,
                        "range": table.ref,
                        "tableStyleInfo": {
                            "name": table.tableStyleInfo.name if table.tableStyleInfo else None,
                            "showFirstColumn": table.tableStyleInfo.showFirstColumn if table.tableStyleInfo else None,
                            "showLastColumn": table.tableStyleInfo.showLastColumn if table.tableStyleInfo else None,
                            "showRowStripes": table.tableStyleInfo.showRowStripes if table.tableStyleInfo else None,
                            "showColumnStripes": table.tableStyleInfo.showColumnStripes if table.tableStyleInfo else None
                        }
                    }
                    
                    # Get column information
                    if hasattr(table, 'tableColumns'):
                        table_data["columns"] = []
                        for col in table.tableColumns:
                            table_data["columns"].append({
                                "name": col.name,
                                "id": col.id
                            })
                    
                    tables.append(table_data)
                    
                except Exception as e:
                    logger.warning(f"Warning: Error processing table {table_name}: {str(e)}")
                    logger.error(traceback.format_exc())
                    continue
                    
        except Exception as e:
            logger.warning(f"Warning: Error extracting tables: {str(e)}")
            logger.error(traceback.format_exc())
            
        return tables

    def _extract_named_ranges(self, sheet) -> List[Dict[str, Any]]:
        """Safely extract named ranges."""
        named_ranges = []
        
        try:
            if not hasattr(self.workbook, 'defined_names') or not self.workbook.defined_names:
                return named_ranges
                
            # Handle both old and new openpyxl versions
            if hasattr(self.workbook.defined_names, 'definedName'):
                # Old openpyxl version
                names = self.workbook.defined_names.definedName
            else:
                # New openpyxl version - defined_names is already a dict-like object
                names = self.workbook.defined_names.values()
                
            for defined_name in names:
                try:
                    # Get the name and value, handling different openpyxl versions
                    name = defined_name.name if hasattr(defined_name, 'name') else defined_name
                    value = defined_name.value if hasattr(defined_name, 'value') else str(defined_name)
                    
                    if sheet.title in str(value):
                        named_range = {
                            "name": name,
                            "value": str(value)
                        }
                        
                        # Add optional fields if they exist
                        if hasattr(defined_name, 'localSheetId'):
                            named_range["localSheetId"] = defined_name.localSheetId
                        if hasattr(defined_name, 'hidden'):
                            named_range["hidden"] = defined_name.hidden
                            
                        named_ranges.append(named_range)
                        
                except Exception as e:
                    logger.warning(f"Warning: Error processing named range: {str(e)}")
                    logger.error(traceback.format_exc())
                    continue
                    
        except Exception as e:
            logger.warning(f"Warning: Error extracting named ranges: {str(e)}")
            logger.error(traceback.format_exc())
            
        return named_ranges

    def _extract_charts_from_sheet(self, sheet) -> List[Dict[str, Any]]:
        """
        Extract comprehensive chart metadata from a worksheet using openpyxl.
        """
        charts = []
        
        try:
            # Get the corresponding sheet from workbook_for_charts (which has read_only=False)
            charts_sheet = None
            if self.workbook_for_charts and sheet.title in self.workbook_for_charts.sheetnames:
                charts_sheet = self.workbook_for_charts[sheet.title]
            
            # If no charts workbook available, fallback to regular sheet
            if not charts_sheet:
                charts_sheet = sheet
            
            # Check if the charts sheet has charts
            if not hasattr(charts_sheet, '_charts') or not charts_sheet._charts:
                return charts
                
            for i, chart in enumerate(charts_sheet._charts):
                try:
                    chart_data = {
                        "chart_name": f"chart{i + 1}",
                        "chart_type": type(chart).__name__,
                        "height": chart.height,
                        "width": chart.width,
                        "style": chart.style,
                        "series_data": {},
                        "series_names": {}
                    }
                    
                    # Extract chart title
                    try:
                        if chart.title and chart.title.tx and chart.title.tx.rich:
                            chart_data["title"] = chart.title.tx.rich.p[0].r[0].t
                        else:
                            chart_data["title"] = None
                    except Exception as e:
                        chart_data["title"] = None
                        logger.debug(f"Could not extract chart title: {str(e)}")
                    
                    # Extract position from anchor
                    try:
                        if hasattr(chart, 'anchor') and chart.anchor:
                            anchor = chart.anchor
                            if hasattr(anchor, '_from'):
                                chart_data["position"] = {
                                    "row": anchor._from.row,
                                    "col": anchor._from.col,
                                    "colOff": anchor._from.colOff,
                                    "rowOff": anchor._from.rowOff
                                }
                                chart_data["left"] = anchor._from.col  # For backwards compatibility
                    except Exception as e:
                        logger.debug(f"Could not extract chart position: {str(e)}")
                    
                    # Extract legend properties
                    try:
                        if hasattr(chart, 'legend') and chart.legend:
                            legend = chart.legend
                            chart_data["legend"] = {
                                "hasLegend": True,
                                "position": getattr(legend, 'position', None),
                                "overlay": getattr(legend, 'overlay', None)
                            }
                        else:
                            chart_data["legend"] = {"hasLegend": False}
                    except Exception as e:
                        chart_data["legend"] = {"hasLegend": False}
                        logger.debug(f"Could not extract legend properties: {str(e)}")
                    
                    # Extract axes properties
                    try:
                        chart_data["axes"] = {}
                        
                        # X-axis
                        if hasattr(chart, 'x_axis') and chart.x_axis:
                            x_axis = chart.x_axis
                            chart_data["axes"]["x_axis"] = {
                                "hasTitle": hasattr(x_axis, 'title') and x_axis.title is not None,
                                "title": None,
                                "hasMajorGridlines": getattr(x_axis, 'majorGridlines', None) is not None,
                                "hasMinorGridlines": getattr(x_axis, 'minorGridlines', None) is not None
                            }
                            
                            # Extract X-axis title
                            try:
                                if x_axis.title and x_axis.title.tx and x_axis.title.tx.rich:
                                    chart_data["axes"]["x_axis"]["title"] = x_axis.title.tx.rich.p[0].r[0].t
                                    chart_data["x_axis"] = x_axis.title.tx.rich.p[0].r[0].t  # For backwards compatibility
                            except:
                                pass
                        
                        # Y-axis
                        if hasattr(chart, 'y_axis') and chart.y_axis:
                            y_axis = chart.y_axis
                            chart_data["axes"]["y_axis"] = {
                                "hasTitle": hasattr(y_axis, 'title') and y_axis.title is not None,
                                "title": None,
                                "hasMajorGridlines": getattr(y_axis, 'majorGridlines', None) is not None,
                                "hasMinorGridlines": getattr(y_axis, 'minorGridlines', None) is not None
                            }
                            
                            # Extract Y-axis title
                            try:
                                if y_axis.title and y_axis.title.tx and y_axis.title.tx.rich:
                                    chart_data["axes"]["y_axis"]["title"] = y_axis.title.tx.rich.p[0].r[0].t
                            except:
                                pass
                    except Exception as e:
                        logger.debug(f"Could not extract axes properties: {str(e)}")
                    
                    # Extract chart background and border properties
                    try:
                        if hasattr(chart, 'graphical_properties') and chart.graphical_properties:
                            gp = chart.graphical_properties
                            chart_data["background"] = {
                                "hasFill": gp.solidFill is not None,
                                "hasGradientFill": gp.gradFill is not None,
                                "noFill": gp.noFill,
                                "hasBorder": gp.ln is not None
                            }
                            
                            # Extract border properties
                            if gp.ln:
                                chart_data["border"] = {
                                    "width": gp.ln.w,
                                    "style": gp.ln.prstDash,
                                    "hasColor": gp.ln.solidFill is not None
                                }
                                
                                # Extract color scheme reference
                                if gp.ln.solidFill and hasattr(gp.ln.solidFill, 'schemeClr'):
                                    chart_data["border"]["colorScheme"] = gp.ln.solidFill.schemeClr.val
                            
                            # Extract background color scheme
                            if gp.solidFill and hasattr(gp.solidFill, 'schemeClr'):
                                chart_data["background"]["colorScheme"] = gp.solidFill.schemeClr.val
                    except Exception as e:
                        logger.debug(f"Could not extract background properties: {str(e)}")
                    
                    # Extract comprehensive series information
                    try:
                        chart_data["series"] = []
                        for j, series in enumerate(chart.series):
                            series_key = f"series_{j + 1}"
                            series_data = {
                                "index": j,
                                "values": series.val.numRef.f if series.val and series.val.numRef else None,
                                "categories": series.cat.strRef.f if series.cat and series.cat.strRef else None,
                                "title": None
                            }
                            
                            # For backwards compatibility
                            chart_data["series_data"][series_key] = series_data["values"]
                            
                            # Extract series title
                            try:
                                if series.tx and series.tx.strRef:
                                    series_data["title"] = series.tx.strRef.f
                                    chart_data["series_names"][f"{series_key}_name"] = series.tx.strRef.f
                                elif series.tx and series.tx.v:
                                    series_data["title"] = series.tx.v
                                    chart_data["series_names"][f"{series_key}_name"] = series.tx.v
                            except:
                                pass
                            
                            # Extract line properties
                            try:
                                if hasattr(series, 'spPr') and series.spPr:
                                    spPr = series.spPr
                                    series_data["line"] = {
                                        "width": spPr.ln.w if spPr.ln else None,
                                        "cap": spPr.ln.cap if spPr.ln else None,
                                        "compound": spPr.ln.cmpd if spPr.ln else None,
                                        "dashStyle": spPr.ln.prstDash if spPr.ln else None,
                                        "hasColor": spPr.ln.solidFill is not None if spPr.ln else False
                                    }
                                    
                                    # Extract color scheme reference
                                    if spPr.ln and spPr.ln.solidFill and hasattr(spPr.ln.solidFill, 'schemeClr'):
                                        series_data["line"]["colorScheme"] = spPr.ln.solidFill.schemeClr.val
                            except Exception as e:
                                logger.debug(f"Could not extract series line properties: {str(e)}")
                            
                            # Extract marker properties
                            try:
                                if hasattr(series, 'marker') and series.marker:
                                    marker = series.marker
                                    series_data["marker"] = {
                                        "hasMarker": True,
                                        "symbol": getattr(marker, 'symbol', None),
                                        "size": getattr(marker, 'size', None),
                                        "hasGraphicalProperties": hasattr(marker, 'spPr') and marker.spPr is not None
                                    }
                                    
                                    # Extract marker graphical properties
                                    if hasattr(marker, 'spPr') and marker.spPr:
                                        markerSpPr = marker.spPr
                                        series_data["marker"]["graphicalProperties"] = {
                                            "hasFill": markerSpPr.solidFill is not None,
                                            "hasLine": markerSpPr.ln is not None
                                        }
                                        
                                        # Extract marker colors
                                        if markerSpPr.solidFill and hasattr(markerSpPr.solidFill, 'schemeClr'):
                                            series_data["marker"]["graphicalProperties"]["fillColorScheme"] = markerSpPr.solidFill.schemeClr.val
                                        if markerSpPr.ln and markerSpPr.ln.solidFill and hasattr(markerSpPr.ln.solidFill, 'schemeClr'):
                                            series_data["marker"]["graphicalProperties"]["lineColorScheme"] = markerSpPr.ln.solidFill.schemeClr.val
                                else:
                                    series_data["marker"] = {"hasMarker": False}
                            except Exception as e:
                                series_data["marker"] = {"hasMarker": False}
                                logger.debug(f"Could not extract marker properties: {str(e)}")
                            
                            # Extract data label properties
                            try:
                                if hasattr(series, 'dLbls') and series.dLbls:
                                    dLbls = series.dLbls
                                    series_data["dataLabels"] = {
                                        "hasDataLabels": True,
                                        "showValue": getattr(dLbls, 'showVal', None),
                                        "showCategoryName": getattr(dLbls, 'showCatName', None),
                                        "showSeriesName": getattr(dLbls, 'showSerName', None),
                                        "showPercent": getattr(dLbls, 'showPercent', None),
                                        "position": getattr(dLbls, 'position', None),
                                        "hasTextProperties": hasattr(dLbls, 'txPr') and dLbls.txPr is not None
                                    }
                                    
                                    # Extract data label font properties
                                    if hasattr(dLbls, 'txPr') and dLbls.txPr:
                                        try:
                                            txPr = dLbls.txPr
                                            if hasattr(txPr, 'p') and txPr.p:
                                                for p in txPr.p:
                                                    if hasattr(p, 'pPr') and p.pPr and hasattr(p.pPr, 'defRPr'):
                                                        defRPr = p.pPr.defRPr
                                                        series_data["dataLabels"]["font"] = {
                                                            "size": defRPr.sz,
                                                            "bold": defRPr.b,
                                                            "italic": defRPr.i,
                                                            "hasColor": defRPr.solidFill is not None
                                                        }
                                                        
                                                        # Extract font family
                                                        if hasattr(defRPr, 'latin') and defRPr.latin:
                                                            series_data["dataLabels"]["font"]["family"] = defRPr.latin.typeface
                                                        
                                                        # Extract font color scheme
                                                        if defRPr.solidFill and hasattr(defRPr.solidFill, 'schemeClr'):
                                                            series_data["dataLabels"]["font"]["colorScheme"] = defRPr.solidFill.schemeClr.val
                                                        break
                                        except Exception as e:
                                            logger.debug(f"Could not extract data label font properties: {str(e)}")
                                else:
                                    series_data["dataLabels"] = {"hasDataLabels": False}
                            except Exception as e:
                                series_data["dataLabels"] = {"hasDataLabels": False}
                                logger.debug(f"Could not extract data label properties: {str(e)}")
                            
                            # Extract error bars
                            try:
                                series_data["errorBars"] = {
                                    "hasErrorBars": hasattr(series, 'errBars') and series.errBars is not None
                                }
                            except Exception as e:
                                series_data["errorBars"] = {"hasErrorBars": False}
                            
                            # Extract trendline
                            try:
                                series_data["trendline"] = {
                                    "hasTrendline": hasattr(series, 'trendline') and series.trendline is not None
                                }
                            except Exception as e:
                                series_data["trendline"] = {"hasTrendline": False}
                            
                            chart_data["series"].append(series_data)
                    except Exception as e:
                        logger.debug(f"Could not extract series information: {str(e)}")
                    
                    # Extract data table information
                    try:
                        chart_data["dataTable"] = {
                            "hasDataTable": hasattr(chart, 'dataTable') and chart.dataTable is not None
                        }
                    except Exception as e:
                        chart_data["dataTable"] = {"hasDataTable": False}
                    
                    charts.append(chart_data)
                    
                except Exception as e:
                    logger.warning(f"Warning: Error processing chart {i}: {str(e)}")
                    logger.error(traceback.format_exc())
                    continue
                    
        except Exception as e:
            logger.warning(f"Warning: Error extracting charts from sheet: {str(e)}")
            logger.error(traceback.format_exc())
            
        return charts

    def _get_theme_colors(self) -> Dict[str, str]:
        """Get theme colors with fallback to defaults."""
        try:
            # Try to get actual theme colors if available
            if hasattr(self.workbook, 'theme') and hasattr(self.workbook.theme, 'theme_elements'):
                # This is a simplified example - actual implementation would need to
                # extract colors from the theme XML
                pass
        except Exception as e:
            logger.warning(f"Warning: Error getting theme colors: {str(e)}")
            logger.error(traceback.format_exc())
        
        # Default theme colors (Office theme) as fallback
        return {
            "background1": "#FFFFFF",
            "background2": "#F2F2F2", 
            "text1": "#000000",
            "text2": "#666666",
            "accent1": "#4472C4",
            "accent2": "#ED7D31",
            "accent3": "#A5A5A5",
            "accent4": "#FFC000",
            "accent5": "#5B9BD5",
            "accent6": "#70AD47",
            "hyperlink": "#0563C1",
            "followedHyperlink": "#954F72"
        }

    def extract_to_json_str(
        self,
        workbook_path: Optional[str] = None,
        output_path: Optional[str] = None,
        max_rows_per_sheet: int = 100,
        max_cols_per_sheet: int = 50,
        include_dependencies: bool = True,
        **kwargs
    ) -> str:
        """
        Extract metadata and return as JSON string with comprehensive error handling and dependencies.
        """
        try:
            # Extract metadata
            metadata, display_values = self.extract_workbook_metadata(
                workbook_path=workbook_path,
                max_rows_per_sheet=max_rows_per_sheet,
                max_cols_per_sheet=max_cols_per_sheet,
                include_dependencies=include_dependencies,
                **kwargs
            )
            
            # Prepare result
            result = {
                "metadata": metadata,
                "displayValues": display_values,
                "status": "success",
                "timestamp": datetime.now().isoformat()
            }
            
            # Convert to JSON
            json_str = json.dumps(result, indent=2, ensure_ascii=False)
            
            # Save to file if output path is provided
            if output_path:
                try:
                    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
                    with open(output_path, 'w', encoding='utf-8') as f:
                        f.write(json_str)
                    logger.info(f"Metadata successfully saved to: {output_path}")
                except Exception as e:
                    logger.error(f"Failed to write output file: {str(e)}")
                    
            return json_str
            
        except Exception as e:
            # Prepare error response
            error_msg = f"Failed to extract metadata to JSON: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            
            error_result = {
                "metadata": None,
                "displayValues": None,
                "status": "error",
                "error": error_msg,
                "timestamp": datetime.now().isoformat()
            }
            
            return json.dumps(error_result, indent=2)


    def _generate_workbook_overview(
        self,
        chunks: List[Dict[str, Any]],
        workbook_path: str
    ) -> Dict[str, Any]:
        """
        Generate a lightweight overview of the workbook structure and content.
        
        Args:
            chunks: List of chunk metadata from extract_workbook_metadata_chunks
            workbook_path: Path to the Excel file
            
        Returns:
            Dictionary containing workbook overview
        """
        overview = {
            "workbook_name": os.path.basename(workbook_path),
            "workbook_path": str(workbook_path),
            "extracted_at": datetime.now().isoformat(),
            "sheets": {}
        }
        
        # Process chunks to build sheet overviews
        for chunk in chunks:
            sheet_name = chunk.get("sheetName")
            if not sheet_name:
                continue
                
            if sheet_name not in overview["sheets"]:
                overview["sheets"][sheet_name] = {
                    "sheet_index": len(overview["sheets"]),
                    "non_empty_rows": set(),
                    "non_empty_columns": set(),
                    "chunks": []
                }
            
            # Update non-empty rows and columns
            sheet_data = overview["sheets"][sheet_name]
            for row_idx, row in enumerate(chunk.get("cellData", []), start=chunk["startRow"]):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.get("value") is not None or cell.get("formula"):
                        sheet_data["non_empty_rows"].add(row_idx)
                        sheet_data["non_empty_columns"].add(col_idx)
            
            # Add chunk info
            chunk_info = {
                "start_row": chunk["startRow"],
                "end_row": chunk["endRow"],
                "row_count": chunk["rowCount"],
                "col_count": chunk["columnCount"],
                "cells": []
            }
            
            # Add cell info (only formula cells)
            for row in chunk.get("cellData", []):
                for cell in row:
                    if cell.get("formula"):
                        chunk_info["cells"].append({
                            "address": cell.get("address"),
                            "formula": cell.get("formula")
                        })
            
            sheet_data["chunks"].append(chunk_info)
        
        # Convert sets to sorted lists
        for sheet in overview["sheets"].values():
            sheet["non_empty_rows"] = sorted(sheet["non_empty_rows"])
            sheet["non_empty_columns"] = sorted(sheet["non_empty_columns"])
        
        return overview

    def extract_workbook_metadata_chunks(
        self,
        workbook_path: Optional[str] = None,
        rows_per_chunk: int = 10,
        max_cols_per_sheet: int = 50,
        include_dependencies: bool = True,
        include_empty_chunks: bool = False,
        force_extract: bool = False
    ) -> List[Dict[str, Any]]:
        """
        Extract metadata in chunks of N rows, returning an array of metadata objects.
        
        Args:
            workbook_path: Path to the Excel file
            rows_per_chunk: Number of rows per chunk (default 10)
            max_cols_per_sheet: Maximum columns to extract per sheet
            include_dependencies: Whether to include dependency analysis
            include_empty_chunks: Whether to include chunks with no data
            force_extract: Whether to force extraction even if storage is available
            
        Returns:
            List of metadata dictionaries, one per chunk
        """
        try:

            normalized_path = str(Path(workbook_path).resolve()) if workbook_path else None
        
            # Check storage first if not forcing extraction
            if self.use_storage and not force_extract and normalized_path:
                logger.info(f"Checking storage for existing metadata for {normalized_path}")
                try:
                    # Get the latest version from storage
                    latest_version = self.storage.get_latest_version(normalized_path)
                    if latest_version:
                        logger.info(f"Latest version found for {normalized_path}: {latest_version}")
                        # Check the file hash:
                        stored_hash = latest_version.get('file_hash')
                        current_hash = self._calculate_file_hash(normalized_path)
                        is_hash_match = stored_hash == current_hash
                        logger.info(f"Hash comparison - Match: {is_hash_match}")                  
                        logger.info(f"Using cached metadata for {normalized_path}")
                        # Get all chunks for this version from db storage
                        chunks = self.storage.get_all_chunks(latest_version['version_id'])
                        
                        if chunks and len(chunks) > 0:
                            logger.info(f"Found {len(chunks)} chunks for {normalized_path}")
                            # check that each chunk is a valid json object
                            if self._are_valid_chunks(chunks):
                                logger.info(f"Chunks validated. Using {len(chunks)} cached metadata chunks for {normalized_path}")
                                return chunks

                            logger.warning(f"Invalid chunks found for {normalized_path}, reverting to forced extraction")
                        logger.warning(f"No chunks found for {normalized_path}, reverting to forced extraction")
                except Exception as e:
                    logger.warning(f"Warning: Error checking storage for existing metadata: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
            if workbook_path or not self.workbook:
                self.open_workbook(workbook_path)
                
            if not self.workbook:
                logger.error("Workbook is not open")

            # Extract metadata chunks from xl file
            logger.info("No cached chunks available. Extracting metadata chunks from xl file...")
            
            chunks = []
            all_cells_metadata = {}  # For dependency analysis across all chunks
            chunk_to_cells_map = {}  # Map chunk index to its cells for dependency updates
            
            # Get workbook info
            workbook_name = os.path.basename(self.workbook_path) if self.workbook_path else "Unknown"
            
            # Process each sheet
            for sheet in self.workbook.worksheets:
                try:
                    sheet_name = sheet.title
                    
                    # Get actual dimensions
                    actual_row_count = min(sheet.max_row or 0, 1048576)
                    actual_col_count = min(sheet.max_column or 0, 16384)
                    
                    if actual_row_count == 0 or actual_col_count == 0:
                        continue
                    
                    # Extract columns to process
                    cols_to_extract = min(actual_col_count, max_cols_per_sheet)
                    
                    # Process sheet in chunks
                    for chunk_start_row in range(1, actual_row_count + 1, rows_per_chunk):
                        chunk_end_row = min(chunk_start_row + rows_per_chunk - 1, actual_row_count)
                        
                        # Extract chunk metadata
                        chunk_metadata = {
                            "chunkId": f"{workbook_name}_{sheet_name}_rows_{chunk_start_row}_{chunk_end_row}",
                            "workbookName": workbook_name,
                            "workbookPath": self.workbook_path,
                            "sheetName": sheet_name,
                            "startRow": chunk_start_row,
                            "endRow": chunk_end_row,
                            "rowCount": chunk_end_row - chunk_start_row + 1,
                            "columnCount": cols_to_extract,
                            "extractedAt": datetime.now().isoformat(),
                            "cellData": [],
                            "chunkIndex": len(chunks),
                            "includeDependencies": include_dependencies
                        }
                        
                        # Extract cell data for this chunk
                        chunk_cells, chunk_cells_dict = self._extract_chunk_cell_data(
                            sheet,
                            chunk_start_row,
                            chunk_end_row,
                            cols_to_extract
                        )
                        
                        # Skip empty chunks if requested
                        if not include_empty_chunks and not any(
                            any(cell.get('value') is not None or cell.get('formula') 
                                for cell in row) 
                            for row in chunk_cells
                        ):
                            continue
                        
                        chunk_metadata["cellData"] = chunk_cells
                        
                        # Store cells for dependency analysis
                        all_cells_metadata.update(chunk_cells_dict)
                        chunk_to_cells_map[len(chunks)] = list(chunk_cells_dict.keys())
                        
                        # Extract tables that intersect with this chunk
                        chunk_metadata["tables"] = self._extract_tables_in_range(
                            sheet,
                            chunk_start_row,
                            chunk_end_row,
                            1,
                            cols_to_extract
                        )
                        
                        chunks.append(chunk_metadata)
                        
                except Exception as e:
                    error_msg = f"Error processing sheet '{sheet.title}': {str(e)}"
                    logger.warning(f"Warning: {error_msg}")
                    logger.error(traceback.format_exc())
                    
                    # Add error chunk
                    chunks.append({
                        "workbookName": workbook_name,
                        "workbookPath": self.workbook_path,
                        "sheetName": sheet.title,
                        "error": error_msg,
                        "chunkIndex": len(chunks)
                    })
            
            # Build dependencies if requested
            if include_dependencies and all_cells_metadata:
                logger.info(f"Building dependencies for {len(chunks)} chunks...")
                try:
                    dependency_map, dependent_map = self.dependency_extractor.build_dependency_maps(all_cells_metadata)
                    
                    # Update each chunk with dependency information
                    for chunk_idx, chunk_cell_addresses in chunk_to_cells_map.items():
                        self._update_chunk_with_dependencies(
                            chunks[chunk_idx],
                            chunk_cell_addresses,
                            dependency_map,
                            dependent_map
                        )
                        
                except Exception as e:
                    logger.warning(f"Warning: Error building dependencies: {str(e)}")
                    logger.error(traceback.format_exc())
            
            # Store the chunks if storage is enabled
            if self.use_storage and normalized_path:
                try:
                    # Create or update workbook in storage
                    logger.info(f"Storing workbook metadata in storage for {normalized_path}")
                    workbook_id = self.storage.create_or_update_workbook(normalized_path)
                    
                    # Create a new version with the extracted chunks
                    new_version_id = self.storage.create_new_version(
                        file_path=normalized_path,
                        change_description=f"Initialized in DB. Chunked extraction with {rows_per_chunk} rows per chunk",
                        chunks=chunks,
                        store_file_blob=True
                    )
                    logger.info(f"Successfully stored chunks in storage for {normalized_path}, new_version_id: {new_version_id}")
                except Exception as e:
                    logger.warning(f"Warning: Failed to store chunks in storage: {str(e)}")
                    logger.error(traceback.format_exc())

            return chunks
            
        except Exception as e:
            error_msg = f"Failed to extract chunk metadata: {str(e)}"
            logger.warning(error_msg)
            logger.error(traceback.format_exc())
            

    def _extract_chunk_cell_data(
        self,
        sheet,
        start_row: int,
        end_row: int,
        col_count: int
    ) -> Tuple[List[List[Dict[str, Any]]], Dict[str, Any]]:
        """
        Extract cell data for a specific chunk range.
        """
        cell_data = []
        cell_dict = {}  # For dependency analysis
        
        try:
            for row_idx in range(start_row, end_row + 1):
                row_data = []
                
                for col_idx in range(1, col_count + 1):
                    try:
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        # Reuse existing method
                        cell_metadata = self._extract_complete_cell_metadata(
                            cell, row_idx, col_idx, sheet.title
                        )
                        row_data.append(cell_metadata)
                        
                        # Add to dictionary for dependency analysis if has content
                        if cell_metadata.get('value') is not None or cell_metadata.get('formula'):
                            full_address = f"{sheet.title}!{cell_metadata['address']}"
                            cell_dict[full_address] = cell_metadata
                            
                    except Exception as e:
                        logger.warning(f"Warning: Error processing cell ({row_idx},{col_idx}): {str(e)}")
                        logger.error(traceback.format_exc())
                        row_data.append({
                            "row": row_idx,
                            "column": col_idx,
                            "address": f"{get_column_letter(col_idx)}{row_idx}",
                            "error": str(e)
                        })
                
                cell_data.append(row_data)
                
        except Exception as e:
            logger.warning(f"Error extracting chunk cell data: {str(e)}")
            logger.error(traceback.format_exc())
                
            
        return cell_data, cell_dict

    def _extract_tables_in_range(
        self,
        sheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int
    ) -> List[Dict[str, Any]]:
        """
        Extract tables that intersect with the given range.
        """
        tables = []
        
        try:
            if not hasattr(sheet, 'tables'):
                return tables
                
            for table_name, table in sheet.tables.items():
                try:
                    # Parse table range
                    table_range = table.ref
                    if ':' in table_range:
                        start_cell, end_cell = table_range.split(':')
                        
                        # Extract row and column from cell addresses
                        table_start_col = self.dependency_extractor.parse_cell_address(start_cell)[0]
                        table_start_row = self.dependency_extractor.parse_cell_address(start_cell)[1]
                        table_end_col = self.dependency_extractor.parse_cell_address(end_cell)[0]
                        table_end_row = self.dependency_extractor.parse_cell_address(end_cell)[1]
                        
                        # Check if table intersects with chunk range
                        if (table_start_row <= end_row and table_end_row >= start_row and
                            table_start_col <= end_col and table_end_col >= start_col):
                            
                            # Add table metadata (reuse existing method logic)
                            table_data = {
                                "name": table_name,
                                "displayName": table.displayName,
                                "range": table.ref,
                                "intersectsChunk": True,
                                "tableStyleInfo": {
                                    "name": table.tableStyleInfo.name if table.tableStyleInfo else None,
                                    "showFirstColumn": table.tableStyleInfo.showFirstColumn if table.tableStyleInfo else None,
                                    "showLastColumn": table.tableStyleInfo.showLastColumn if table.tableStyleInfo else None,
                                    "showRowStripes": table.tableStyleInfo.showRowStripes if table.tableStyleInfo else None,
                                    "showColumnStripes": table.tableStyleInfo.showColumnStripes if table.tableStyleInfo else None
                                }
                            }
                            
                            # Get column information if available
                            if hasattr(table, 'tableColumns'):
                                table_data["columns"] = []
                                for col in table.tableColumns:
                                    table_data["columns"].append({
                                        "name": col.name,
                                        "id": col.id
                                    })
                            
                            tables.append(table_data)
                            
                except Exception as e:
                    logger.warning(f"Warning: Error processing table {table_name} for chunk: {str(e)}")
                    logger.error(traceback.format_exc())
                    continue
                    
        except Exception as e:
            logger.warning(f"Warning: Error extracting tables for chunk: {str(e)}")
            logger.error(traceback.format_exc())
            
        return tables

    def _update_chunk_with_dependencies(
        self,
        chunk_metadata: Dict[str, Any],
        chunk_cell_addresses: List[str],
        dependency_map: Dict[str, set],
        dependent_map: Dict[str, set]
    ):
        """
        Update chunk cells with dependency information.
        """
        # Update cells in the chunk
        for row_data in chunk_metadata["cellData"]:
            for cell_metadata in row_data:
                if "address" in cell_metadata:
                    full_address = f"{chunk_metadata['sheetName']}!{cell_metadata['address']}"
                    
                    # Get dependency information
                    precedents = dependency_map.get(full_address, set())
                    dependents = dependent_map.get(full_address, set())
                    
                    # Update cell metadata (reuse existing logic)
                    cell_metadata.update({
                        "directPrecedents": list(precedents),
                        "directDependents": list(dependents),
                        "precedentCount": len(precedents),
                        "dependentCount": len(dependents),
                        "totalConnections": len(precedents) + len(dependents)
                    })
        
        # Add chunk-level dependency summary
        chunk_precedents = set()
        chunk_dependents = set()
        internal_dependencies = 0
        external_dependencies = 0
        
        for cell_addr in chunk_cell_addresses:
            # Collect all dependencies for this chunk
            cell_precedents = dependency_map.get(cell_addr, set())
            cell_dependents = dependent_map.get(cell_addr, set())
            
            chunk_precedents.update(cell_precedents)
            chunk_dependents.update(cell_dependents)
            
            # Count internal vs external dependencies
            for prec in cell_precedents:
                if prec in chunk_cell_addresses:
                    internal_dependencies += 1
                else:
                    external_dependencies += 1
        
        # Add dependency summary to chunk
        chunk_metadata["dependencySummary"] = {
            "totalPrecedents": len(chunk_precedents),
            "totalDependents": len(chunk_dependents),
            "internalDependencies": internal_dependencies,
            "externalDependencies": external_dependencies,
            "externalPrecedents": list(chunk_precedents - set(chunk_cell_addresses))[:10],  # Top 10
            "externalDependents": list(chunk_dependents - set(chunk_cell_addresses))[:10]   # Top 10
        }


    def extract_lightweight_metadata(self, workbook_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Extract lightweight metadata containing just cell addresses, formulas, and values.
        Returns data in a format compatible with hotcache.
        
        Args:
            workbook_path: Path to the Excel file (optional if already opened)
            
        Returns:
            Dict containing workbook info and cell data in hotcache-compatible format
        """
        try:
            if workbook_path or not self.workbook:
                self.open_workbook(workbook_path)
                
            if not self.workbook:
                logger.error("Workbook is not open")
                
            # Basic workbook info
            workbook_info = {
                "workbook_name": os.path.basename(self.workbook_path) if self.workbook_path else "Unknown",
                "sheet_names": [sheet.title for sheet in self.workbook.worksheets],
                "total_sheets": len(self.workbook.worksheets),
                "sheets": {}
            }
            
            # Process each sheet
            for sheet in self.workbook.worksheets:
                sheet_name = sheet.title
                workbook_info["sheets"][sheet_name] = {
                    "sheet_index": len(workbook_info["sheets"]),
                    "chunks": [{
                        "startRow": 1,
                        "endRow": min(sheet.max_row or 0, 1048576),
                        "rowCount": min(sheet.max_row or 0, 1048576),
                        "columnCount": min(sheet.max_column or 0, 16384),
                        "chunkIndex": 0,
                        "cells": self._extract_sheet_cells_lightweight(sheet),
                        "charts": self._extract_charts_from_sheet(sheet)
                    }]
                }
                
            return workbook_info
            
        except Exception as e:
            error_msg = f"Error extracting lightweight metadata: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
        finally:
            self.close()

    def _extract_sheet_cells_lightweight(self, sheet) -> List[Dict[str, Any]]:
        """
        Extract lightweight cell data including address, formula, value, and formatting properties.
        
        Args:
            sheet: OpenPyXL worksheet object
            
        Returns:
            List of cell data dictionaries
        """
        cells = []
        
        try:
            # Get actual dimensions
            max_row = min(sheet.max_row or 0, 1048576)
            max_col = min(sheet.max_column or 0, 16384)
            
            if max_row == 0 or max_col == 0:
                return []
                
            # Get corresponding sheet from values workbook
            value_sheet = None
            if self.workbook_values and sheet.title in self.workbook_values.sheetnames:
                value_sheet = self.workbook_values[sheet.title]
            
            # Extract cell data
            for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    try:
                        # Skip empty cells
                        if cell.value is None and not (hasattr(cell, 'value') and cell.data_type == 'f'):
                            continue
                            
                        # Get cell value
                        cell_value = None
                        if value_sheet:
                            try:
                                value_cell = value_sheet.cell(row=cell.row, column=cell.column)
                                cell_value = self._serialize_value(value_cell.value)
                            except:
                                pass
                        
                        # Get formula if present
                        formula = None
                        if hasattr(cell, 'data_type') and cell.data_type == 'f' and cell.value is not None:
                            formula = str(cell.value).lstrip('=')
                        
                        # Extract formatting properties
                        formatting = self._extract_complete_cell_formatting(cell)
                        
                        # Only include cells with values or formulas
                        if cell_value is not None or formula:
                            cell_data = {
                                "a": f"{get_column_letter(cell.column)}{cell.row}",  # address
                                "v": cell_value,  # value
                                "f": formula  # formula
                            }
                            
                            # Add formatting properties if they exist
                            if formatting:
                                cell_data["fmt"] = formatting
                            
                            cells.append(cell_data)
                            
                    except Exception as e:
                        logger.warning(f"Error processing cell {cell.coordinate}: {str(e)}")
                        continue
                        
        except Exception as e:
            logger.error(f"Error extracting cells from sheet {sheet.title}: {str(e)}")
            logger.error(traceback.format_exc())

            
        return cells


    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate a hash of the file contents."""
        if not os.path.exists(file_path):
            return ""
            
        try:
            with open(file_path, 'rb') as f:
                return hashlib.sha256(f.read()).hexdigest()
        except Exception:
            return ""


    def _is_valid_metadata(self, metadata: Dict[str, Any]) -> bool:
        """
        Validate that the metadata has the expected structure and contains data.
        
        Args:
            metadata: The metadata dictionary to validate
            
        Returns:
            bool: True if metadata is valid, False otherwise
        """
        if not isinstance(metadata, dict):
            return False
        
        # Check for required top-level keys
        required_keys = {'file_info', 'sheets'}
        if not required_keys.issubset(metadata.keys()):
            return False
        
        # Check that sheets is a non-empty list
        if not isinstance(metadata.get('sheets'), list) or not metadata['sheets']:
            return False
        
        # Check that at least one sheet has content
        return any(
            isinstance(sheet, dict) and 
            sheet.get('cells') and 
            isinstance(sheet['cells'], list)
            for sheet in metadata['sheets']
        )


    def _are_valid_chunks(self, chunks: List[Dict[str, Any]]) -> bool:
        """Validate that all chunks in the list are valid."""
        return all(self._is_valid_chunk(chunk) for chunk in chunks)


    def _is_valid_chunk(self, chunk: dict) -> bool:
        """Validate that a chunk has the required structure and data."""
        if not chunk or not isinstance(chunk, dict):
            return False
            
        # Check for required top-level fields
        required_fields = {'sheetName', 'startRow', 'endRow', 'cellData'}
        if not all(field in chunk for field in required_fields):
            return False
            
        # Check cellData is a non-empty 2D array
        cell_data = chunk.get('cellData', [])
        if not isinstance(cell_data, list) or not cell_data:
            return False
        if not all(isinstance(row, list) for row in cell_data):
            return False
        if not any(len(row) > 0 for row in cell_data):  # At least one non-empty row
            return False
            
        # Check row bounds make sense
        if not (0 <= chunk['startRow'] <= chunk['endRow']):
            return False
            
        return True