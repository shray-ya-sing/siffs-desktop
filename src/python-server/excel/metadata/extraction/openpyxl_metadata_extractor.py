import openpyxl
from openpyxl.styles import Font, Fill, Border, Alignment
from openpyxl.utils import get_column_letter
import json
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import os

class OpenpyxlExcelMetadataExtractor:
    """
    Python class for extracting comprehensive cell-based metadata from Excel workbooks using openpyxl.
    Focuses on data and formatting that openpyxl handles extremely well.
    """
    def __init__(self, workbook_path: Optional[str] = None):
        """
        Initialize the metadata extractor.
        
        Args:
            workbook_path: Path to the Excel file (optional, can be set later)
        """
        self.workbook_path = workbook_path
        self.workbook = None
        self.workbook_values = None  # Add this line
    
    def open_workbook(self, workbook_path: Optional[str] = None) -> None:
        """
        Open the Excel workbook with both formula and value access.
        
        Args:
            workbook_path: Path to the Excel file
        """
        if workbook_path:
            self.workbook_path = workbook_path
            
        if not self.workbook_path:
            raise ValueError("No workbook path specified")
            
        if not os.path.exists(self.workbook_path):
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")
            
        try:
            # Open workbook twice - once for formulas, once for values
            self.workbook = openpyxl.load_workbook(self.workbook_path, data_only=False)  # For formulas
            self.workbook_values = openpyxl.load_workbook(self.workbook_path, data_only=True)  # For calculated values
            
        except Exception as e:
            raise Exception(f"Failed to open workbook: {str(e)}")

    def close(self) -> None:
        """Close both workbooks"""
        try:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
            if hasattr(self, 'workbook_values') and self.workbook_values:
                self.workbook_values.close()
                self.workbook_values = None
        except Exception as e:
            print(f"Warning: Error during cleanup: {str(e)}")
        
    def __enter__(self):
        """Context manager entry"""
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - cleanup resources"""
        self.close()
        
                
    def extract_workbook_metadata(self, workbook_path: Optional[str] = None, max_rows_per_sheet: int = 100, max_cols_per_sheet: int = 50) -> Dict[str, Any]:
        """
        Extract comprehensive metadata from the Excel workbook with cell-based structure.
        
        Args:
            workbook_path: Path to the Excel file
            max_rows_per_sheet: Maximum number of rows to extract per sheet
            max_cols_per_sheet: Maximum number of columns to extract per sheet
            
        Returns:
            Dictionary containing complete workbook metadata
        """
        if workbook_path or not self.workbook:
            self.open_workbook(workbook_path)
            
        try:
            # Get basic workbook info
            workbook_metadata = {
                "extractedAt": datetime.now().isoformat(),
                "workbookPath": self.workbook_path,
                "workbookName": os.path.basename(self.workbook_path),
                "activeSheet": self.workbook.active.title,
                "totalSheets": len(self.workbook.worksheets),
                "sheetNames": [sheet.title for sheet in self.workbook.worksheets],
                "themeColors": self._get_theme_colors(),
                "sheets": []
            }
            
            # Extract metadata for each sheet
            for sheet in self.workbook.worksheets:
                sheet_metadata = self._extract_sheet_metadata(sheet, max_rows_per_sheet, max_cols_per_sheet)
                workbook_metadata["sheets"].append(sheet_metadata)
                
            return workbook_metadata
            
        except Exception as e:
            raise Exception(f"Failed to extract workbook metadata: {str(e)}")
            
    def _extract_sheet_metadata(self, sheet, max_rows: int, max_cols: int) -> Dict[str, Any]:
        """
        Extract comprehensive metadata from a single sheet with cell-based structure.
        
        Args:
            sheet: openpyxl Worksheet object
            max_rows: Maximum number of rows to extract
            max_cols: Maximum number of columns to extract
            
        Returns:
            Dictionary containing sheet metadata
        """
        try:
            # Get actual dimensions
            actual_row_count = sheet.max_row
            actual_col_count = sheet.max_column
            
            # Check if sheet is empty
            if actual_row_count == 1 and actual_col_count == 1 and sheet.cell(1, 1).value is None:
                return {
                    "name": sheet.title,
                    "isEmpty": True,
                    "rowCount": 0,
                    "columnCount": 0,
                    "cellData": [],
                    "tables": [],
                    "namedRanges": []
                }
            
            # Determine extraction bounds
            extract_row_count = min(actual_row_count, max_rows)
            extract_col_count = min(actual_col_count, max_cols)
            
            # Extract cell data in the required format
            cell_data = self._extract_cell_data(sheet, extract_row_count, extract_col_count)
            
            # Extract other sheet elements
            tables = self._extract_tables_metadata(sheet)
            named_ranges = self._extract_named_ranges(sheet)
            
            return {
                "name": sheet.title,
                "isEmpty": False,
                "rowCount": actual_row_count,
                "columnCount": actual_col_count,
                "extractedRowCount": extract_row_count,
                "extractedColumnCount": extract_col_count,
                "cellData": cell_data,
                "tables": tables,
                "namedRanges": named_ranges
            }
            
        except Exception as e:
            print(f"Warning: Error extracting metadata for sheet '{sheet.title}': {str(e)}")
            return {
                "name": sheet.title,
                "error": str(e),
                "isEmpty": True,
                "rowCount": 0,
                "columnCount": 0,
                "cellData": [],
                "tables": [],
                "namedRanges": []
            }
            
    def _extract_cell_data(self, sheet, row_count: int, col_count: int) -> List[List[Dict[str, Any]]]:
        """
        Extract comprehensive cell data including values, formulas, and all formatting.
        
        Args:
            sheet: openpyxl Worksheet object
            row_count: Number of rows to extract
            col_count: Number of columns to extract
            
        Returns:
            2D array of cell objects with complete metadata
        """
        try:
            cell_data = []
            
            # Extract data row by row
            for row_idx in range(1, row_count + 1):
                row_data = []
                
                for col_idx in range(1, col_count + 1):
                    # Get cell
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    # Extract complete cell metadata
                    cell_metadata = self._extract_complete_cell_metadata(cell, row_idx, col_idx)
                    row_data.append(cell_metadata)
                    
                cell_data.append(row_data)
                
            return cell_data
            
        except Exception as e:
            print(f"Warning: Error extracting cell data: {str(e)}")
            return []
        
    def _extract_complete_cell_metadata(self, cell, row: int, col: int) -> Dict[str, Any]:
        """
        Extract complete metadata for a single cell including all possible properties.
        
        Args:
            cell: openpyxl Cell object (from formula workbook)
            row: 1-based row number
            col: 1-based column number
            
        Returns:
            Dictionary containing all cell metadata
        """
        try:
            # Get corresponding cell from values workbook using sheet name
            value_sheet = self.workbook_values[cell.parent.title]  # Use bracket notation instead
            value_cell = value_sheet.cell(row=row, column=col)
            
            cell_metadata = {
                "row": row,
                "column": col,
                "address": f"{get_column_letter(col)}{row}",
                "value": self._serialize_value(value_cell.value),  # Calculated value
                "formula": self._get_cell_formula(cell),  # Formula from formula workbook
                "formatting": self._extract_complete_cell_formatting(cell)
            }
            
            return cell_metadata
            
        except Exception as e:
            print(f"Warning: Error extracting cell metadata for {row},{col}: {str(e)}")
            return {
                "row": row,
                "column": col,
                "address": f"{get_column_letter(col)}{row}",
                "value": None,
                "formula": None,
                "formatting": {},
                "error": str(e)
            }
        
        
    def _get_cell_formula(self, cell) -> Optional[str]:
        """Get the formula from a cell (from formula workbook)"""
        try:
            # Check if cell has a formula
            if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                return cell.value
            return None
        except:
            return None

            
    def _extract_complete_cell_formatting(self, cell) -> Dict[str, Any]:
        """
        Extract all possible formatting properties from a cell using openpyxl.
        
        Args:
            cell: openpyxl Cell object
            
        Returns:
            Dictionary containing all formatting properties
        """
        formatting = {}
        
        try:
            # Font properties
            try:
                if cell.font:
                    formatting["font"] = {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "color": self._get_color_hex(cell.font.color),
                        "underline": cell.font.underline,
                        "strikethrough": cell.font.strikethrough,
                        "vertAlign": cell.font.vertAlign  # subscript/superscript
                    }
            except Exception as e:
                formatting["font"] = {"error": str(e)}
                
            # Fill/Background properties
            try:
                if cell.fill:
                    formatting["fill"] = {
                        "fillType": cell.fill.fill_type,
                        "startColor": self._get_color_hex(cell.fill.start_color),
                        "endColor": self._get_color_hex(cell.fill.end_color),
                        "patternType": cell.fill.patternType
                    }
            except Exception as e:
                formatting["fill"] = {"error": str(e)}
                
            # Number format
            try:
                formatting["numberFormat"] = cell.number_format
            except Exception as e:
                formatting["numberFormat"] = {"error": str(e)}
                
            # Alignment properties
            try:
                if cell.alignment:
                    formatting["alignment"] = {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                        "wrapText": cell.alignment.wrap_text,
                        "shrinkToFit": cell.alignment.shrink_to_fit,
                        "indent": cell.alignment.indent,
                        "textRotation": cell.alignment.text_rotation,
                        "readingOrder": cell.alignment.reading_order,
                        "justifyLastLine": cell.alignment.justify_last_line,
                        "relativeIndent": cell.alignment.relative_indent
                    }
            except Exception as e:
                formatting["alignment"] = {"error": str(e)}
                
            # Border properties
            try:
                if cell.border:
                    formatting["borders"] = {
                        "left": self._get_border_info(cell.border.left),
                        "right": self._get_border_info(cell.border.right),
                        "top": self._get_border_info(cell.border.top),
                        "bottom": self._get_border_info(cell.border.bottom),
                        "diagonal": self._get_border_info(cell.border.diagonal),
                        "diagonalUp": cell.border.diagonal_up,
                        "diagonalDown": cell.border.diagonal_down,
                        "outline": cell.border.outline,
                        "start": self._get_border_info(cell.border.start),
                        "end": self._get_border_info(cell.border.end)
                    }
            except Exception as e:
                formatting["borders"] = {"error": str(e)}
                
            # Protection properties
            try:
                if cell.protection:
                    formatting["protection"] = {
                        "locked": cell.protection.locked,
                        "hidden": cell.protection.hidden
                    }
            except Exception as e:
                formatting["protection"] = {"error": str(e)}
                
            # Comments/Notes
            try:
                if cell.comment:
                    formatting["comment"] = {
                        "text": cell.comment.text,
                        "author": cell.comment.author,
                        "width": cell.comment.width,
                        "height": cell.comment.height
                    }
            except Exception as e:
                formatting["comment"] = {"error": str(e)}
                
            # Hyperlink
            try:
                if cell.hyperlink:
                    formatting["hyperlink"] = {
                        "target": cell.hyperlink.target,
                        "tooltip": cell.hyperlink.tooltip,
                        "display": cell.hyperlink.display
                    }
            except Exception as e:
                formatting["hyperlink"] = {"error": str(e)}
                
            # Data type
            try:
                formatting["dataType"] = cell.data_type
            except Exception as e:
                formatting["dataType"] = {"error": str(e)}
                
            # Merged cell information
            try:
                merged_ranges = cell.parent.merged_cells.ranges
                is_merged = False
                merge_range = None
                
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        merge_range = str(merged_range)
                        break
                        
                formatting["merged"] = {
                    "isMerged": is_merged,
                    "mergeRange": merge_range
                }
            except Exception as e:
                formatting["merged"] = {"error": str(e)}
                
        except Exception as e:
            formatting["extractionError"] = str(e)
            
        return formatting
        
    def _get_color_hex(self, color) -> Optional[str]:
        """Convert openpyxl color to hex string"""
        try:
            if not color:
                return None
                
            # Handle RGB color
            if hasattr(color, 'rgb') and color.rgb:
                rgb = color.rgb
                if len(rgb) == 8:  # ARGB format
                    return f"#{rgb[2:]}"  # Remove alpha channel
                elif len(rgb) == 6:  # RGB format
                    return f"#{rgb}"
                    
            # Handle indexed color
            if hasattr(color, 'indexed') and color.indexed is not None:
                return f"indexed_{color.indexed}"
                
            # Handle theme color
            if hasattr(color, 'theme') and color.theme is not None:
                return f"theme_{color.theme}"
                
            # Handle auto color
            if hasattr(color, 'auto') and color.auto:
                return "auto"
                
            return str(color)
        except:
            return None
            
    def _get_border_info(self, border_side) -> Dict[str, Any]:
        """Get border information for a specific side"""
        try:
            if not border_side:
                return {"style": None, "color": None}
                
            return {
                "style": border_side.style,
                "color": self._get_color_hex(border_side.color)
            }
        except:
            return {"style": None, "color": None}
            
    def _serialize_value(self, value: Any) -> Any:
        """Serialize a cell value to JSON-compatible format"""
        if value is None:
            return None
        elif isinstance(value, (str, int, float, bool)):
            return value
        elif hasattr(value, 'isoformat'):  # datetime objects
            return value.isoformat()
        else:
            return str(value)
            
    def _extract_tables_metadata(self, sheet) -> List[Dict[str, Any]]:
        """Extract metadata about tables in the sheet"""
        try:
            tables = []
            
            # Get tables from the sheet
            if hasattr(sheet, 'tables'):
                for table_name, table in sheet.tables.items():
                    try:
                        table_data = {
                            "name": table_name,
                            "displayName": table.displayName,
                            "range": table.ref,
                            "tableStyleInfo": {
                                "name": table.tableStyleInfo.name if table.tableStyleInfo else None,
                                "showFirstColumn": table.tableStyleInfo.showFirstColumn if table.tableStyleInfo else None,
                                "showLastColumn": table.tableStyleInfo.showLastColumn if table.tableStyleInfo else None,
                                "showRowStripes": table.tableStyleInfo.showRowStripes if table.tableStyleInfo else None,
                                "showColumnStripes": table.tableStyleInfo.showColumnStripes if table.tableStyleInfo else None
                            }
                        }
                        
                        # Get column information
                        if table.tableColumns:
                            table_data["columns"] = []
                            for col in table.tableColumns:
                                table_data["columns"].append({
                                    "name": col.name,
                                    "id": col.id
                                })
                        
                        tables.append(table_data)
                        
                    except Exception as e:
                        print(f"Warning: Error processing table {table_name}: {str(e)}")
                        continue
                        
            return tables
            
        except Exception as e:
            print(f"Warning: Error extracting tables metadata: {str(e)}")
            return []
            
    def _extract_named_ranges(self, sheet) -> List[Dict[str, Any]]:
        """Extract named ranges that reference this sheet"""
        try:
            named_ranges = []
            
            # Get named ranges from workbook
            if hasattr(self.workbook, 'defined_names'):
                for defined_name in self.workbook.defined_names.definedName:
                    try:
                        # Check if the named range refers to this sheet
                        if sheet.title in str(defined_name.value):
                            named_ranges.append({
                                "name": defined_name.name,
                                "value": str(defined_name.value),
                                "localSheetId": defined_name.localSheetId,
                                "hidden": defined_name.hidden
                            })
                    except Exception as e:
                        print(f"Warning: Error processing named range: {str(e)}")
                        continue
                        
            return named_ranges
            
        except Exception as e:
            print(f"Warning: Error extracting named ranges: {str(e)}")
            return []
            
    def _get_theme_colors(self) -> Dict[str, str]:
        """Get theme colors from the workbook"""
        try:
            # Default theme colors (Office theme)
            return {
                "background1": "#FFFFFF",
                "background2": "#F2F2F2", 
                "text1": "#000000",
                "text2": "#666666",
                "accent1": "#4472C4",
                "accent2": "#ED7D31",
                "accent3": "#A5A5A5",
                "accent4": "#FFC000",
                "accent5": "#5B9BD5",
                "accent6": "#70AD47",
                "hyperlink": "#0563C1",
                "followedHyperlink": "#954F72"
            }
        except Exception as e:
            print(f"Warning: Error getting theme colors: {str(e)}")
            return {}
            
    def extract_to_json(self, workbook_path: Optional[str] = None, output_path: Optional[str] = None, 
                       max_rows_per_sheet: int = 100, max_cols_per_sheet: int = 50) -> str:
        """
        Extract metadata and return as JSON string.
        
        Args:
            workbook_path: Path to the Excel file
            output_path: Optional path to save JSON file
            max_rows_per_sheet: Maximum number of rows to extract per sheet
            max_cols_per_sheet: Maximum number of columns to extract per sheet
            
        Returns:
            JSON string containing the metadata
        """
        metadata = self.extract_workbook_metadata(workbook_path, max_rows_per_sheet, max_cols_per_sheet)
        json_str = json.dumps(metadata, indent=2, ensure_ascii=False)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(json_str)
            print(f"Metadata saved to: {output_path}")
            
        return json_str

    def extract_json_compress(self, workbook_path: Optional[str] = None, output_path: Optional[str] = None, max_rows_per_sheet: int = 100, max_cols_per_sheet: int = 50) -> str:
        """
        Extract metadata and return as compressed JSON string with descriptive field names.
        
        Args:
            workbook_path: Path to the Excel file
            output_path: Optional path to save JSON file
            max_rows_per_sheet: Maximum number of rows to extract per sheet
            max_cols_per_sheet: Maximum number of columns to extract per sheet
            
        Returns:
            Compressed JSON string with full field names for readability
        """
        metadata = self.extract_workbook_metadata(workbook_path, max_rows_per_sheet, max_cols_per_sheet)
        
        # Compress the metadata structure
        compressed = self._compress_metadata(metadata)
        
        # Convert to JSON with minimal formatting
        json_str = json.dumps(compressed, separators=(',', ':'), ensure_ascii=False)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(json_str)
            print(f"Metadata saved to: {output_path}")
        
        return json_str

    def _compress_metadata(self, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Compress metadata structure to minimize tokens while keeping descriptive field names.
        
        Args:
            metadata: Full metadata dictionary
            
        Returns:
            Compressed metadata with full field names
        """
        compressed = {
            "workbookName": metadata.get("workbookName", ""),
            "activeSheet": metadata.get("activeSheet", ""),
            "totalSheets": metadata.get("totalSheets", 0),
            "sheetNames": metadata.get("sheetNames", []),
            "sheets": []
        }
        
        for sheet in metadata.get("sheets", []):
            compressed_sheet = {
                "name": sheet.get("name", ""),
                "rowCount": sheet.get("rowCount", 0),
                "columnCount": sheet.get("columnCount", 0),
                "extractedRowCount": sheet.get("extractedRowCount", 0),
                "extractedColumnCount": sheet.get("extractedColumnCount", 0),
                "cells": self._compress_cells(sheet.get("cellData", [])),
                "tables": self._compress_tables(sheet.get("tables", [])),
                "namedRanges": self._compress_named_ranges(sheet.get("namedRanges", []))
            }
            compressed["sheets"].append(compressed_sheet)
        
        return compressed

    def _compress_cells(self, cell_data: List[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        """
        Compress cell data to flat list with descriptive field names.
        Only include cells with meaningful data.
        
        Args:
            cell_data: 2D array of cell objects
            
        Returns:
            Flattened list of compressed cell objects
        """
        compressed_cells = []
        
        for row_data in cell_data:
            for cell in row_data:
                # Only include cells with value, formula, or significant formatting
                if self._is_significant_cell(cell):
                    compressed_cell = {
                        "row": cell.get("row"),
                        "column": cell.get("column")
                    }
                    
                    # Add value if present
                    value = cell.get("value")
                    if value is not None and value != "":
                        compressed_cell["value"] = value
                    
                    # Add formula if present
                    formula = cell.get("formula")
                    if formula and formula != str(value):
                        compressed_cell["formula"] = formula
                    
                    # Add compressed formatting if significant
                    formatting = self._compress_cell_formatting(cell.get("formatting", {}))
                    if formatting:
                        compressed_cell["formatting"] = formatting
                    
                    compressed_cells.append(compressed_cell)
        
        return compressed_cells

    def _compress_cell_formatting(self, formatting: Dict[str, Any]) -> Dict[str, Any]:
        """
        Compress cell formatting to minimal representation with descriptive field names.
        
        Args:
            formatting: Full formatting dictionary
            
        Returns:
            Compressed formatting dictionary
        """
        compressed = {}
        
        try:
            # Font properties (only if non-default)
            font = formatting.get("font", {})
            font_compressed = {}
            
            if font.get("bold"):
                font_compressed["isBold"] = True
            if font.get("italic"):
                font_compressed["isItalic"] = True
            if font.get("color"):
                font_compressed["fontColor"] = font["color"]
            if font.get("size") and font["size"] != 11:  # Default size
                font_compressed["fontSize"] = font["size"]
            if font.get("name") and font["name"] != "Calibri":  # Default font
                font_compressed["fontName"] = font["name"]
            if font.get("underline"):
                font_compressed["underline"] = font["underline"]
            if font.get("strikethrough"):
                font_compressed["strikethrough"] = True
            
            if font_compressed:
                compressed["font"] = font_compressed
            
            # Fill color
            fill = formatting.get("fill", {})
            if fill.get("startColor"):
                compressed["backgroundColor"] = fill["startColor"]
            
            # Number format (only if not general)
            number_format = formatting.get("numberFormat")
            if number_format and number_format.lower() not in ["general", "@"]:
                compressed["numberFormat"] = number_format
            
            # Alignment (only if not default)
            alignment = formatting.get("alignment", {})
            if alignment.get("horizontal"):
                compressed["horizontalAlignment"] = alignment["horizontal"]
            if alignment.get("vertical"):
                compressed["verticalAlignment"] = alignment["vertical"]
            if alignment.get("wrapText"):
                compressed["wrapText"] = True
            
            # Borders (simplified)
            borders = formatting.get("borders", {})
            border_styles = []
            for side, border in borders.items():
                if isinstance(border, dict) and border.get("style"):
                    border_styles.append(side)
            if border_styles:
                compressed["hasBorders"] = border_styles
            
            # Merged cells
            merged = formatting.get("merged", {})
            if merged.get("isMerged"):
                compressed["isMerged"] = True
                if merged.get("mergeRange"):
                    compressed["mergeRange"] = merged["mergeRange"]
            
            # Hyperlink
            hyperlink = formatting.get("hyperlink", {})
            if hyperlink.get("target"):
                compressed["hyperlink"] = hyperlink["target"]
            
            # Comment
            comment = formatting.get("comment", {})
            if comment.get("text"):
                compressed["comment"] = comment["text"][:100]  # Limit comment length
            
        except Exception as e:
            pass
        
        return compressed

    def _compress_tables(self, tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Compress table metadata with descriptive field names"""
        compressed_tables = []
        
        for table in tables:
            compressed_table = {
                "tableName": table.get("name", ""),
                "displayName": table.get("displayName", ""),
                "tableRange": table.get("range", ""),
                "columns": [col.get("name", "") for col in table.get("columns", [])][:10]  # Limit columns
            }
            compressed_tables.append(compressed_table)
        
        return compressed_tables

    def _compress_named_ranges(self, named_ranges: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Compress named ranges metadata with descriptive field names"""
        compressed_ranges = []
        
        for nr in named_ranges:
            compressed_range = {
                "rangeName": nr.get("name", ""),
                "rangeValue": nr.get("value", ""),
                "hidden": nr.get("hidden", False)
            }
            compressed_ranges.append(compressed_range)
        
        return compressed_ranges
    
    def extract_markdown_compress(self, workbook_path: Optional[str] = None, output_path: Optional[str] = None, max_rows_per_sheet: int = 100, max_cols_per_sheet: int = 50) -> str:
        """
        Extract metadata and return as compact markdown table format for LLM.
        
        Args:
            workbook_path: Path to the Excel file
            output_path: Optional path to save markdown file
            max_rows_per_sheet: Maximum number of rows to extract per sheet
            max_cols_per_sheet: Maximum number of columns to extract per sheet
            
        Returns:
            Markdown string with cell data in table format
        """
        metadata = self.extract_workbook_metadata(workbook_path, max_rows_per_sheet, max_cols_per_sheet)
        
        markdown_lines = []
        markdown_lines.append(f"# Workbook: {metadata.get('workbookName', '')}")
        markdown_lines.append(f"Active Sheet: {metadata.get('activeSheet', '')}")
        markdown_lines.append(f"Total Sheets: {metadata.get('totalSheets', 0)}")
        markdown_lines.append("")
        
        for sheet in metadata.get("sheets", []):
            if sheet.get("isEmpty", True):
                continue
                
            markdown_lines.append(f"## Sheet: {sheet.get('name', '')}")
            markdown_lines.append(f"Dimensions: {sheet.get('rowCount', 0)} rows x {sheet.get('columnCount', 0)} columns")
            markdown_lines.append("")
            
            # Process cell data
            significant_cells = []
            for row_data in sheet.get("cellData", []):
                for cell in row_data:
                    if self._is_significant_cell(cell):
                        significant_cells.append(cell)
            
            if significant_cells:
                # Create markdown table header
                markdown_lines.append("| address | value | formula | fill_color | font_color | font_bold | font_italic | font_size | font_name | number_format | h_align | v_align | wrap_text | borders | merged | hyperlink | comment |")
                markdown_lines.append("|---------|-------|---------|------------|------------|-----------|-------------|-----------|-----------|---------------|---------|---------|-----------|---------|--------|-----------|---------|")
                
                # Add cell data rows
                for cell in significant_cells:
                    row = cell.get("row", "")
                    col = cell.get("column", "")
                    address = f"R{row}C{col}"
                    
                    value = cell.get("value", "")
                    if value is None:
                        value = ""
                    
                    formula = cell.get("formula", "")
                    if formula:
                        formula = f'"{formula}"'
                    else:
                        formula = ""
                    
                    formatting = cell.get("formatting", {})
                    
                    # Extract formatting details
                    fill_color = self._get_fill_color_compact(formatting)
                    font_info = self._get_font_info_compact(formatting)
                    alignment_info = self._get_alignment_info_compact(formatting)
                    border_info = self._get_border_info_compact(formatting)
                    other_info = self._get_other_info_compact(formatting)
                    
                    # Create table row - ensure all values are strings
                    row_data = [
                        str(address),
                        str(value),
                        str(formula),
                        str(fill_color or ""),
                        str(font_info.get("color") or ""),
                        "Y" if font_info.get("bold") else "",
                        "Y" if font_info.get("italic") else "",
                        str(font_info.get("size") or ""),
                        str(font_info.get("name") or ""),
                        str(formatting.get("numberFormat") or ""),
                        str(alignment_info.get("horizontal") or ""),
                        str(alignment_info.get("vertical") or ""),
                        "Y" if alignment_info.get("wrap") else "",
                        str(border_info or ""),
                        "Y" if other_info.get("merged") else "",
                        str(other_info.get("hyperlink") or ""),
                        str(other_info.get("comment") or "")
                    ]
                    
                    markdown_lines.append("| " + " | ".join(row_data) + " |")
            
            # Add tables info
            tables = sheet.get("tables", [])
            if tables:
                markdown_lines.append("")
                markdown_lines.append("### Tables:")
                for table in tables:
                    table_name = table.get("tableName", table.get("name", ""))
                    table_range = table.get("tableRange", table.get("range", ""))
                    headers = table.get("headers", [])
                    markdown_lines.append(f"- **{table_name}** ({table_range}): {', '.join(headers[:5])}")
            
            # Add named ranges info
            named_ranges = sheet.get("namedRanges", [])
            if named_ranges:
                markdown_lines.append("")
                markdown_lines.append("### Named Ranges:")
                for nr in named_ranges:
                    range_name = nr.get("rangeName", nr.get("name", ""))
                    range_ref = nr.get("rangeReference", nr.get("value", ""))
                    markdown_lines.append(f"- **{range_name}**: {range_ref}")
            
            markdown_lines.append("")
        
        markdown_string = "\n".join(markdown_lines)        
    
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_string)
            print(f"Metadata saved to: {output_path}")
        
        return markdown_string

    def _get_fill_color_compact(self, formatting: Dict[str, Any]) -> str:
        """Extract fill color in compact format"""
        try:
            fill = formatting.get("fill", {})
            start_color = fill.get("startColor", "")
            if start_color and start_color != "auto":
                return start_color
            return ""
        except:
            return ""

    def _get_font_info_compact(self, formatting: Dict[str, Any]) -> Dict[str, Any]:
        """Extract font info in compact format"""
        try:
            font = formatting.get("font", {})
            return {
                "color": font.get("color", ""),
                "bold": font.get("bold", False),
                "italic": font.get("italic", False),
                "size": font.get("size", ""),
                "name": font.get("name", "")
            }
        except:
            return {"color": "", "bold": False, "italic": False, "size": "", "name": ""}

    def _get_alignment_info_compact(self, formatting: Dict[str, Any]) -> Dict[str, Any]:
        """Extract alignment info in compact format"""
        try:
            alignment = formatting.get("alignment", {})
            return {
                "horizontal": alignment.get("horizontal", ""),
                "vertical": alignment.get("vertical", ""),
                "wrap": alignment.get("wrapText", False)
            }
        except:
            return {"horizontal": "", "vertical": "", "wrap": False}

    def _get_border_info_compact(self, formatting: Dict[str, Any]) -> str:
        """Extract border info in compact format"""
        try:
            borders = formatting.get("borders", {})
            border_sides = []
            for side, border in borders.items():
                if isinstance(border, dict) and border.get("style"):
                    border_sides.append(side[0].upper())  # First letter uppercase
            return "".join(sorted(border_sides))
        except:
            return ""

    def _get_other_info_compact(self, formatting: Dict[str, Any]) -> Dict[str, Any]:
        """Extract other formatting info in compact format"""
        try:
            merged = formatting.get("merged", {})
            hyperlink = formatting.get("hyperlink", {})
            comment = formatting.get("comment", {})
            
            return {
                "merged": merged.get("isMerged", False),
                "hyperlink": hyperlink.get("target", "")[:30] if hyperlink.get("target") else "",
                "comment": comment.get("text", "")[:30] if comment.get("text") else ""
            }
        except:
            return {"merged": False, "hyperlink": "", "comment": ""}
    
    def extract_markdown_spreadsheet_style(self, workbook_path: Optional[str] = None, output_path: Optional[str] = None, max_rows_per_sheet: int = 100, max_cols_per_sheet: int = 50) -> str:
        """
        Extract metadata and return as spreadsheet-style markdown for LLM analysis.
        
        Args:
            workbook_path: Path to the Excel file
            output_path: Optional path to save markdown file
            max_rows_per_sheet: Maximum number of rows to extract per sheet
            max_cols_per_sheet: Maximum number of columns to extract per sheet
            
        Returns:
            Markdown string with spreadsheet-style layout
        """
        metadata = self.extract_workbook_metadata(workbook_path, max_rows_per_sheet, max_cols_per_sheet)
        
        markdown_lines = []
        markdown_lines.append(f"# Workbook: {metadata.get('workbookName', '')}")
        markdown_lines.append(f"Active Sheet: {metadata.get('activeSheet', '')}")
        markdown_lines.append(f"Total Sheets: {metadata.get('totalSheets', 0)}")
        markdown_lines.append("")
        
        for sheet in metadata.get("sheets", []):
            if sheet.get("isEmpty", True):
                continue
                
            markdown_lines.append(f"## Sheet: {sheet.get('name', '')}")
            markdown_lines.append(f"Dimensions: {sheet.get('rowCount', 0)} rows x {sheet.get('columnCount', 0)} columns")
            markdown_lines.append("")
            
            # Create spreadsheet-style table
            cell_data = sheet.get("cellData", [])
            if cell_data:
                # Create header row with column letters
                max_cols = len(cell_data[0]) if cell_data else 0
                header_row = ["Row"] + [get_column_letter(i+1) for i in range(max_cols)]
                markdown_lines.append("| " + " | ".join(header_row) + " |")
                markdown_lines.append("|" + "---|" * len(header_row))
                
                # Add data rows
                for row_idx, row_data in enumerate(cell_data):
                    row_cells = [str(row_idx + 1)]  # Row number
                    
                    for cell in row_data:
                        cell_content = self._format_cell_for_spreadsheet(cell)
                        row_cells.append(cell_content)
                    
                    markdown_lines.append("| " + " | ".join(row_cells) + " |")
            
            # Add tables info
            tables = sheet.get("tables", [])
            if tables:
                markdown_lines.append("")
                markdown_lines.append("### Tables:")
                for table in tables:
                    table_name = table.get("tableName", table.get("name", ""))
                    table_range = table.get("tableRange", table.get("range", ""))
                    headers = table.get("headers", [])
                    markdown_lines.append(f"- **{table_name}** ({table_range}): {', '.join(headers[:5])}")
            
            # Add named ranges info
            named_ranges = sheet.get("namedRanges", [])
            if named_ranges:
                markdown_lines.append("")
                markdown_lines.append("### Named Ranges:")
                for nr in named_ranges:
                    range_name = nr.get("rangeName", nr.get("name", ""))
                    range_ref = nr.get("rangeReference", nr.get("value", ""))
                    markdown_lines.append(f"- **{range_name}**: {range_ref}")
            
            markdown_lines.append("")
        
        markdown_string = "\n".join(markdown_lines)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_string)
            print(f"Metadata saved to: {output_path}")

        return markdown_string
    
    def extract_markdown_spreadsheet_style_comprehensively(self, workbook_path: Optional[str] = None, output_path: Optional[str] = None, max_rows_per_sheet: int = 100, max_cols_per_sheet: int = 50) -> str:
        """
        Extract metadata and return as spreadsheet-style markdown for LLM analysis.
        
        Args:
            workbook_path: Path to the Excel file
            output_path: Optional path to save markdown file
            max_rows_per_sheet: Maximum number of rows to extract per sheet
            max_cols_per_sheet: Maximum number of columns to extract per sheet
            
        Returns:
            Markdown string with spreadsheet-style layout
        """
        metadata = self.extract_workbook_metadata(workbook_path, max_rows_per_sheet, max_cols_per_sheet)
        
        markdown_lines = []
        markdown_lines.append(f"# Workbook: {metadata.get('workbookName', '')}")
        markdown_lines.append(f"Active Sheet: {metadata.get('activeSheet', '')}")
        markdown_lines.append(f"Total Sheets: {metadata.get('totalSheets', 0)}")
        markdown_lines.append("")
        
        for sheet in metadata.get("sheets", []):
            if sheet.get("isEmpty", True):
                continue
                
            markdown_lines.append(f"## Sheet: {sheet.get('name', '')}")
            markdown_lines.append(f"Dimensions: {sheet.get('rowCount', 0)} rows x {sheet.get('columnCount', 0)} columns")
            markdown_lines.append("")
            
            # Create spreadsheet-style table
            cell_data = sheet.get("cellData", [])
            if cell_data:
                # Create header row with column letters
                max_cols = len(cell_data[0]) if cell_data else 0
                header_row = ["Row"] + [get_column_letter(i+1) for i in range(max_cols)]
                markdown_lines.append("| " + " | ".join(header_row) + " |")
                markdown_lines.append("|" + "---|" * len(header_row))
                
                # Add data rows
                for row_idx, row_data in enumerate(cell_data):
                    row_cells = [str(row_idx + 1)]  # Row number
                    
                    for cell in row_data:
                        cell_content = self._format_cell_for_spreadsheet_complete(cell)
                        row_cells.append(cell_content)
                    
                    markdown_lines.append("| " + " | ".join(row_cells) + " |")
            
            # Add tables info
            tables = sheet.get("tables", [])
            if tables:
                markdown_lines.append("")
                markdown_lines.append("### Tables:")
                for table in tables:
                    table_name = table.get("tableName", table.get("name", ""))
                    table_range = table.get("tableRange", table.get("range", ""))
                    headers = table.get("headers", [])
                    display_name = table.get("displayName", "")
                    style_info = table.get("tableStyleInfo", {})
                    
                    table_details = f"- **{table_name}** ({table_range})"
                    if display_name and display_name != table_name:
                        table_details += f" Display: {display_name}"
                    if headers:
                        table_details += f" Headers: {', '.join(headers[:10])}"
                    if style_info.get("name"):
                        table_details += f" Style: {style_info['name']}"
                    
                    markdown_lines.append(table_details)
            
            # Add named ranges info
            named_ranges = sheet.get("namedRanges", [])
            if named_ranges:
                markdown_lines.append("")
                markdown_lines.append("### Named Ranges:")
                for nr in named_ranges:
                    range_name = nr.get("rangeName", nr.get("name", ""))
                    range_ref = nr.get("rangeReference", nr.get("value", ""))
                    hidden = nr.get("hidden", False)
                    local_sheet = nr.get("localSheetId", "")
                    
                    range_details = f"- **{range_name}**: {range_ref}"
                    if hidden:
                        range_details += " (Hidden)"
                    if local_sheet:
                        range_details += f" (Sheet-specific: {local_sheet})"
                    
                    markdown_lines.append(range_details)
            
            markdown_lines.append("")
        
        markdown_string = "\n".join(markdown_lines)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_string)
            print(f"Metadata saved to: {output_path}")

        return markdown_string

    def _format_cell_for_spreadsheet(self, cell: Dict[str, Any]) -> str:
        """
        Format cell data with ALL available properties for spreadsheet-style display.
        
        Args:
            cell: Cell metadata dictionary
            
        Returns:
            Formatted string with ALL cell properties
        """
        try:
            properties = []
            
            # Value (always include, even if empty)
            value = cell.get("value", "")
            if value is not None and str(value).strip():
                properties.append(f"value={str(value)}")
            else:
                properties.append("value=null")
            
            # Formula (wrapped in quotes)
            formula = cell.get("formula", "")
            if formula:
                properties.append(f'formula="{formula}"')
            
            formatting = cell.get("formatting", {})
            
            # Font properties - ALL of them
            font = formatting.get("font", {})
            if font.get("name") and font.get("name") != "Calibri":
                properties.append(f"font_name={font['name']}")
            if font.get("size") and font.get("size") != 11:
                properties.append(f"font_size={font['size']}")
            if font.get("bold"):
                properties.append("bold=Y")
            if font.get("italic"):
                properties.append("italic=Y")
            if font.get("color"):
                properties.append(f"font_color={font['color']}")
            if font.get("underline") and font.get("underline") != "none":
                properties.append(f"underline={font['underline']}")
            if font.get("strikethrough"):
                properties.append("strikethrough=Y")
            if font.get("vertAlign"):
                properties.append(f"vert_align={font['vertAlign']}")
            
            # Fill properties - ALL of them
            fill = formatting.get("fill", {})
            if fill.get("fillType") and fill.get("fillType") != "none":
                properties.append(f"fill_type={fill['fillType']}")
            if fill.get("startColor"):
                properties.append(f"bg_color={fill['startColor']}")
            if fill.get("endColor") and fill.get("endColor") != fill.get("startColor"):
                properties.append(f"end_color={fill['endColor']}")
            if fill.get("patternType"):
                properties.append(f"pattern={fill['patternType']}")
            
            # Number format
            number_format = formatting.get("numberFormat", "")
            if number_format and number_format.lower() not in ["general", "@"]:
                properties.append(f"format={number_format}")
            
            # Alignment properties - ALL of them
            alignment = formatting.get("alignment", {})
            if alignment.get("horizontal"):
                properties.append(f"h_align={alignment['horizontal']}")
            if alignment.get("vertical"):
                properties.append(f"v_align={alignment['vertical']}")
            if alignment.get("wrapText"):
                properties.append("wrap=Y")
            if alignment.get("shrinkToFit"):
                properties.append("shrink=Y")
            if alignment.get("indent") and alignment.get("indent") != 0:
                properties.append(f"indent={alignment['indent']}")
            if alignment.get("textRotation") and alignment.get("textRotation") != 0:
                properties.append(f"rotation={alignment['textRotation']}")
            if alignment.get("readingOrder") and alignment.get("readingOrder") != 0:
                properties.append(f"reading_order={alignment['readingOrder']}")
            if alignment.get("justifyLastLine"):
                properties.append("justify_last=Y")
            if alignment.get("relativeIndent") and alignment.get("relativeIndent") != 0:
                properties.append(f"rel_indent={alignment['relativeIndent']}")
            
            # Border properties - ALL of them
            borders = formatting.get("borders", {})
            border_details = []
            for side in ["left", "right", "top", "bottom", "diagonal", "start", "end"]:
                border = borders.get(side, {})
                if isinstance(border, dict) and border.get("style"):
                    border_info = f"{side[0].upper()}:{border['style']}"
                    if border.get("color"):
                        border_info += f"({border['color']})"
                    border_details.append(border_info)
            
            if border_details:
                properties.append(f"borders={','.join(border_details)}")
            
            if borders.get("diagonalUp"):
                properties.append("diag_up=Y")
            if borders.get("diagonalDown"):
                properties.append("diag_down=Y")
            if borders.get("outline"):
                properties.append("outline=Y")
            
            # Protection properties
            protection = formatting.get("protection", {})
            if protection.get("locked") is False:  # Only show if explicitly unlocked
                properties.append("locked=N")
            if protection.get("hidden"):
                properties.append("hidden=Y")
            
            # Data type
            data_type = formatting.get("dataType", "")
            if data_type and data_type != "n":  # n is default for numbers
                properties.append(f"data_type={data_type}")
            
            # Merged cell information
            merged = formatting.get("merged", {})
            if merged.get("isMerged"):
                properties.append("merged=Y")
                if merged.get("mergeRange"):
                    properties.append(f"merge_range={merged['mergeRange']}")
            
            # Hyperlink
            hyperlink = formatting.get("hyperlink", {})
            if hyperlink.get("target"):
                properties.append(f"link={hyperlink['target'][:30]}")
                if hyperlink.get("tooltip"):
                    properties.append(f"tooltip={hyperlink['tooltip'][:20]}")
                if hyperlink.get("display"):
                    properties.append(f"display={hyperlink['display'][:20]}")
            
            # Comment
            comment = formatting.get("comment", {})
            if comment.get("text"):
                properties.append(f"note={comment['text'][:30]}")
                if comment.get("author"):
                    properties.append(f"note_author={comment['author']}")
                if comment.get("width") or comment.get("height"):
                    properties.append(f"note_size={comment.get('width', 0)}x{comment.get('height', 0)}")
            
            # Always return properties (even for empty cells with formatting)
            return ", ".join(properties)
                
        except Exception as e:
            return f"ERROR: {str(e)[:30]}"

    def _is_significant_cell(self, cell: Dict[str, Any]) -> bool:
        """
        Determine if a cell has significant data worth including.
        Modified to include empty cells with formatting.
        
        Args:
            cell: Cell metadata dictionary
            
        Returns:
            True if cell should be included in compressed output
        """
        value = cell.get("value")
        formula = cell.get("formula")
        formatting = cell.get("formatting", {})
        
        # Include if has non-empty value
        if value is not None and str(value).strip() != "":
            return True
        
        # Include if has formula
        if formula and formula.startswith("="):
            return True
        
        # Include if has significant formatting (even if empty)
        if self._has_significant_formatting(formatting):
            return True
        
        return False

    def _has_significant_formatting(self, formatting: Dict[str, Any]) -> bool:
        """Check if formatting is significant enough to include - comprehensive version"""
        try:
            # Check font formatting
            font = formatting.get("font", {})
            if (font.get("bold") or font.get("italic") or font.get("color") or 
                font.get("underline") != "none" or font.get("strikethrough") or
                font.get("vertAlign") or (font.get("size") and font.get("size") != 11) or
                (font.get("name") and font.get("name") != "Calibri")):
                return True
            
            # Check fill formatting
            fill = formatting.get("fill", {})
            if (fill.get("startColor") or fill.get("endColor") or 
                fill.get("fillType") != "none" or fill.get("patternType")):
                return True
            
            # Check borders
            borders = formatting.get("borders", {})
            for border in borders.values():
                if isinstance(border, dict) and border.get("style"):
                    return True
            if borders.get("diagonalUp") or borders.get("diagonalDown") or borders.get("outline"):
                return True
            
            # Check alignment
            alignment = formatting.get("alignment", {})
            if (alignment.get("horizontal") or alignment.get("vertical") or
                alignment.get("wrapText") or alignment.get("shrinkToFit") or
                alignment.get("indent") != 0 or alignment.get("textRotation") != 0 or
                alignment.get("readingOrder") != 0 or alignment.get("justifyLastLine") or
                alignment.get("relativeIndent") != 0):
                return True
            
            # Check number format (if not general)
            number_format = formatting.get("numberFormat")
            if number_format and number_format.lower() not in ["general", "@"]:
                return True
            
            # Check protection
            protection = formatting.get("protection", {})
            if protection.get("locked") is False or protection.get("hidden"):
                return True
            
            # Check data type
            data_type = formatting.get("dataType", "")
            if data_type and data_type != "n":
                return True
            
            # Check merged cells
            merged = formatting.get("merged", {})
            if merged.get("isMerged"):
                return True
            
            # Check hyperlink
            if formatting.get("hyperlink"):
                return True
            
            # Check comment
            if formatting.get("comment"):
                return True
            
            return False
        except:
            return False
    
    def _format_cell_for_spreadsheet(self, cell: Dict[str, Any]) -> str:
        """
        Format cell data as comma-separated properties for spreadsheet-style display.
        
        Args:
            cell: Cell metadata dictionary
            
        Returns:
            Formatted string with cell properties
        """
        try:
            properties = []
            
            # Value (always include, even if empty)
            value = cell.get("value", "")
            if value is not None and str(value).strip():
                properties.append(f"value={str(value)}")
            else:
                properties.append("value=null")
            
            # Formula (wrapped in quotes)
            formula = cell.get("formula", "")
            if formula:
                properties.append(f'formula="{formula}"')
            
            formatting = cell.get("formatting", {})
            
            # Font properties
            font = formatting.get("font", {})
            if font.get("bold"):
                properties.append("bold=Y")
            if font.get("italic"):
                properties.append("italic=Y")
            if font.get("color"):
                properties.append(f"font_color={font['color']}")
            if font.get("size") and font.get("size") != 11:
                properties.append(f"font_size={font['size']}")
            if font.get("name") and font.get("name") != "Calibri":
                properties.append(f"font={font['name']}")
            
            # Fill color
            fill = formatting.get("fill", {})
            if fill.get("startColor"):
                properties.append(f"bg_color={fill['startColor']}")
            
            # Number format
            number_format = formatting.get("numberFormat", "")
            if number_format and number_format.lower() not in ["general", "@"]:
                properties.append(f"format={number_format}")
            
            # Alignment
            alignment = formatting.get("alignment", {})
            if alignment.get("horizontal"):
                properties.append(f"h_align={alignment['horizontal']}")
            if alignment.get("vertical"):
                properties.append(f"v_align={alignment['vertical']}")
            if alignment.get("wrapText"):
                properties.append("wrap=Y")
            
            # Borders
            borders = formatting.get("borders", {})
            border_sides = []
            for side, border in borders.items():
                if isinstance(border, dict) and border.get("style"):
                    border_sides.append(side[0].upper())
            if border_sides:
                properties.append(f"borders={''.join(sorted(border_sides))}")
            
            # Other properties
            merged = formatting.get("merged", {})
            if merged.get("isMerged"):
                properties.append("merged=Y")
            
            hyperlink = formatting.get("hyperlink", {})
            if hyperlink.get("target"):
                properties.append(f"link={hyperlink['target'][:20]}")
            
            comment = formatting.get("comment", {})
            if comment.get("text"):
                properties.append(f"note={comment['text'][:20]}")
            
            # Always return properties (even for empty cells with formatting)
            return ", ".join(properties)
                
        except Exception as e:
            return f"ERROR: {str(e)[:20]}"

    
