from typing import Tuple, Dict, Any, Optional
from openpyxl.utils import get_column_letter

class SpreadsheetMarkdownCompressor:
    """
    Python class for compressing json metadata into spreadsheet style markdown with dependency information
    """
    def __init__(self, workbook_path: Optional[str] = None):
        """
        Initialize the markdown compressor.        
        """
        self.has_display_values = False
        self.has_dependencies = False

    def compress_to_markdown(self, metadata_tuple: Tuple[Dict[str, Any], Optional[Dict[str, str]]], output_path: Optional[str] = r'C:\Users\shrey\OneDrive\Desktop\cori outputs\appTest.txt') -> str:
        """
        Compress the metadata passed as arg into spreadsheet-style markdown for LLM analysis.
        
        Args:
            metadata_tuple: Tuple of the OpenPyxl metadata and optional Xlwings extracted display values                
            output_path: Optional path to save markdown file
        
        Returns:
            Markdown string with spreadsheet-style layout including dependency information
        """
        # Unpack metadata and display values (which might be None)
        if isinstance(metadata_tuple, (list, tuple)) and len(metadata_tuple) >= 2:
            metadata, display_values = metadata_tuple
            self.has_display_values = bool(display_values)
        else:
            # Handle case where only metadata is provided
            metadata = metadata_tuple if not isinstance(metadata_tuple, (list, tuple)) else metadata_tuple[0]
            display_values = {}
            self.has_display_values = False

        # Check if dependencies are included
        self.has_dependencies = metadata.get('includeDependencies', False)

        markdown_lines = []
        markdown_lines.append(f"# Workbook: {metadata.get('workbookName', '')}")
        markdown_lines.append(f"Active Sheet: {metadata.get('activeSheet', '')}")
        markdown_lines.append(f"Total Sheets: {metadata.get('totalSheets', 0)}")
        
        # Add dependency summary if available
        if self.has_dependencies and 'dependencySummary' in metadata:
            self._add_dependency_summary_section(markdown_lines, metadata['dependencySummary'])
        
        markdown_lines.append("")
        
        for sheet in metadata.get("sheets", []):
            if sheet.get("isEmpty", True):
                continue
                
            markdown_lines.append(f"## Sheet: {sheet.get('name', '')}")
            markdown_lines.append(f"Dimensions: {sheet.get('rowCount', 0)} rows x {sheet.get('columnCount', 0)} columns")
            
            # Add sheet-level dependency statistics
            if self.has_dependencies:
                self._add_sheet_dependency_stats(markdown_lines, sheet)
            
            markdown_lines.append("")
            
            # Create spreadsheet-style table
            cell_data = sheet.get("cellData", [])
            if cell_data:
                # Create header row with column letters
                max_cols = len(cell_data[0]) if cell_data else 0
                header_row = ["Row"] + [get_column_letter(i+1) for i in range(max_cols)]
                markdown_lines.append("| " + " | ".join(header_row) + " |")
                markdown_lines.append("|" + "---|" * len(header_row))
                
                # Add data rows
                for row_idx, row_data in enumerate(cell_data):
                    row_cells = [str(row_idx + 1)]  # Row number
                    
                    for cell in row_data:
                        # Get display value from xlwings data if available
                        display_value = None
                        if self.has_display_values and display_values:
                            sheet_name = sheet.get('name', '')
                            cell_key = f"{sheet_name}_{cell.get('row')}_{cell.get('column')}"
                            display_value = display_values.get(cell_key)
                        
                        cell_content = self._format_cell_for_spreadsheet(cell, display_value)
                        row_cells.append(cell_content)
                    
                    markdown_lines.append("| " + " | ".join(row_cells) + " |")
            
            # Add dependency highlights for this sheet
            if self.has_dependencies:
                self._add_sheet_dependency_highlights(markdown_lines, sheet)
            
            # Rest of the existing functionality
            tables = sheet.get("tables", [])
            if tables:
                markdown_lines.append("")
                markdown_lines.append("### Tables:")
                for table in tables:
                    table_name = table.get("tableName", table.get("name", ""))
                    table_range = table.get("tableRange", table.get("range", ""))
                    headers = table.get("headers", [])
                    markdown_lines.append(f"- **{table_name}** ({table_range}): {', '.join(headers[:5])}")
            
            named_ranges = sheet.get("namedRanges", [])
            if named_ranges:
                markdown_lines.append("")
                markdown_lines.append("### Named Ranges:")
                for nr in named_ranges:
                    range_name = nr.get("rangeName", nr.get("name", ""))
                    range_ref = nr.get("rangeReference", nr.get("value", ""))
                    markdown_lines.append(f"- **{range_name}**: {range_ref}")
            
            markdown_lines.append("")
        
        markdown_string = "\n".join(markdown_lines)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_string)
            print(f"Metadata saved to: {output_path}")

        return markdown_string

    def _add_dependency_summary_section(self, markdown_lines: list, dependency_summary: Dict[str, Any]):
        """Add workbook-level dependency summary section"""
        markdown_lines.append("")
        markdown_lines.append("## Dependency Analysis Summary")
        markdown_lines.append(f"- **Total Cells:** {dependency_summary.get('totalCells', 0):,}")
        markdown_lines.append(f"- **Formula Cells:** {dependency_summary.get('formulaCells', 0):,}")
        markdown_lines.append(f"- **Value Cells:** {dependency_summary.get('valueCells', 0):,}")
        markdown_lines.append(f"- **Total Dependencies:** {dependency_summary.get('totalDependencies', 0):,}")
        
        avg_deps = dependency_summary.get('avgDependenciesPerCell', 0)
        markdown_lines.append(f"- **Average Dependencies per Cell:** {avg_deps:.2f}")
        
        # Most connected cells
        most_connected = dependency_summary.get('mostConnectedCells', [])[:5]
        if most_connected:
            markdown_lines.append("")
            markdown_lines.append("### Most Connected Cells:")
            for item in most_connected:
                cell_addr = item.get('cell', '')
                connections = item.get('connections', 0)
                markdown_lines.append(f"- **{cell_addr}**: {connections} total connections")
        
        # Most complex formulas
        most_complex = dependency_summary.get('mostComplexFormulas', [])[:5]
        if most_complex:
            markdown_lines.append("")
            markdown_lines.append("### Most Complex Formulas:")
            for item in most_complex:
                cell_addr = item.get('cell', '')
                precedents = item.get('precedents', 0)
                markdown_lines.append(f"- **{cell_addr}**: {precedents} precedents")
        
        # Most referenced cells
        most_referenced = dependency_summary.get('mostReferencedCells', [])[:5]
        if most_referenced:
            markdown_lines.append("")
            markdown_lines.append("### Most Referenced Cells:")
            for item in most_referenced:
                cell_addr = item.get('cell', '')
                dependents = item.get('dependents', 0)
                markdown_lines.append(f"- **{cell_addr}**: {dependents} dependents")

    def _add_sheet_dependency_stats(self, markdown_lines: list, sheet: Dict[str, Any]):
        """Add sheet-level dependency statistics"""
        cell_data = sheet.get("cellData", [])
        if not cell_data:
            return
        
        # Calculate sheet-level stats
        total_cells_with_deps = 0
        total_precedents = 0
        total_dependents = 0
        max_precedents = 0
        max_dependents = 0
        
        for row_data in cell_data:
            for cell in row_data:
                precedent_count = cell.get('precedentCount', 0)
                dependent_count = cell.get('dependentCount', 0)
                
                if precedent_count > 0 or dependent_count > 0:
                    total_cells_with_deps += 1
                    total_precedents += precedent_count
                    total_dependents += dependent_count
                    max_precedents = max(max_precedents, precedent_count)
                    max_dependents = max(max_dependents, dependent_count)
        
        if total_cells_with_deps > 0:
            markdown_lines.append(f"**Dependencies:** {total_cells_with_deps} cells with connections, ")
            markdown_lines.append(f"{total_precedents} total precedents, {total_dependents} total dependents")

    def _add_sheet_dependency_highlights(self, markdown_lines: list, sheet: Dict[str, Any]):
        """Add highlights of interesting dependency patterns in this sheet"""
        cell_data = sheet.get("cellData", [])
        if not cell_data:
            return
        
        # Find cells with high connectivity
        high_precedent_cells = []
        high_dependent_cells = []
        formula_chains = []
        
        for row_data in cell_data:
            for cell in row_data:
                precedent_count = cell.get('precedentCount', 0)
                dependent_count = cell.get('dependentCount', 0)
                address = cell.get('address', '')
                formula = cell.get('formula', '')
                
                # High precedent cells (complex formulas)
                if precedent_count >= 5:
                    high_precedent_cells.append({
                        'address': address,
                        'precedents': precedent_count,
                        'formula': formula[:50] + '...' if len(formula) > 50 else formula
                    })
                
                # High dependent cells (key inputs)
                if dependent_count >= 3:
                    high_dependent_cells.append({
                        'address': address,
                        'dependents': dependent_count,
                        'value': str(cell.get('value', ''))[:30]
                    })
                
                # Formula chains (cells that both depend on others and are depended upon)
                if precedent_count >= 2 and dependent_count >= 2:
                    formula_chains.append({
                        'address': address,
                        'precedents': precedent_count,
                        'dependents': dependent_count
                    })
        
        # Sort and limit results
        high_precedent_cells.sort(key=lambda x: x['precedents'], reverse=True)
        high_dependent_cells.sort(key=lambda x: x['dependents'], reverse=True)
        formula_chains.sort(key=lambda x: x['precedents'] + x['dependents'], reverse=True)
        
        # Add to markdown if any interesting patterns found
        if high_precedent_cells[:3] or high_dependent_cells[:3] or formula_chains[:3]:
            markdown_lines.append("")
            markdown_lines.append("### Dependency Highlights:")
            
            if high_precedent_cells[:3]:
                markdown_lines.append("**Complex Formulas:**")
                for cell in high_precedent_cells[:3]:
                    markdown_lines.append(f"- {cell['address']}: {cell['precedents']} precedents")
                    if cell['formula']:
                        markdown_lines.append(f"  Formula: `{cell['formula']}`")
            
            if high_dependent_cells[:3]:
                markdown_lines.append("**Key Input Cells:**")
                for cell in high_dependent_cells[:3]:
                    markdown_lines.append(f"- {cell['address']}: {cell['dependents']} dependents")
                    if cell['value']:
                        markdown_lines.append(f"  Value: {cell['value']}")
            
            if formula_chains[:3]:
                markdown_lines.append("**Calculation Chain Nodes:**")
                for cell in formula_chains[:3]:
                    markdown_lines.append(f"- {cell['address']}: {cell['precedents']} → {cell['dependents']}")

    def _format_cell_for_spreadsheet(self, cell: Dict[str, Any], display_value: Optional[str] = None) -> str:
        """
        Format cell data with critical available properties for spreadsheet-style display,
        now including dependency information.
        
        Args:
            cell: Cell metadata dictionary
            display_value: Optional display value from xlwings (what user sees)
            
        Returns:
            Formatted string with critical cell properties including dependencies
        """
        def _safe_format_value(value: Any) -> str:
            """
            Safely format values for markdown without escaping - use quotes when needed.
            
            Args:
                value: Value to format
                
            Returns:
                Safely formatted value
            """
            if value is None:
                return 'null'
            if not isinstance(value, str):
                value = str(value)
            
            # Characters that might cause issues in markdown tables
            problematic_chars = ['#', '*', '_', '`', '[', ']', '(', ')', '|', '$', '^', '~', '\\', '"', "'"]
            
            # If value contains problematic characters, wrap in quotes
            if any(char in value for char in problematic_chars):
                # Escape any existing quotes in the value
                safe_value = value.replace('"', "'")  # Replace double quotes with single quotes
                return f'"{safe_value}"'
            
            return value

        def _format_dependency_list(dep_list: list, max_items: int = 3) -> str:
            """Format a list of dependencies for compact display"""
            if not dep_list:
                return "none"
            
            # Show first few items, indicate if there are more
            displayed = dep_list[:max_items]
            result = ",".join(displayed)
            
            if len(dep_list) > max_items:
                result += f",+{len(dep_list) - max_items}more"
            
            return result

        try:
            properties = []
            
            # Cell location information (always first)
            row = cell.get("row", "")
            col = cell.get("column", "")
            address = cell.get("address", f"{get_column_letter(col)}{row}" if col and row else "")
            
            if address:
                properties.append(f'addr={_safe_format_value(address)}')
            
            # Raw value (from openpyxl)
            raw_value = cell.get("value", "")
            if raw_value is not None and str(raw_value).strip():
                properties.append(f'val={_safe_format_value(str(raw_value))}')
            else:
                properties.append('val=null')
            
            # Display value (from xlwings) - only include if provided and different
            if display_value is not None and str(display_value) != str(raw_value):
                properties.append(f'disp={_safe_format_value(display_value)}')
            
            # Formula (if present)
            formula = cell.get("formula")
            if formula and formula.startswith("="):
                # Truncate long formulas
                formula_display = formula[:30] + "..." if len(formula) > 30 else formula
                properties.append(f'form={_safe_format_value(formula_display)}')
            
            # Dependency information (if dependencies are included)
            if self.has_dependencies:
                precedent_count = cell.get("precedentCount", 0)
                dependent_count = cell.get("dependentCount", 0)
                total_connections = cell.get("totalConnections", 0)
                
                # Include counts if there are any connections
                if total_connections > 0:
                    properties.append(f'deps={precedent_count}→{dependent_count}')
                    
                    # Include actual precedents/dependents for high-value cells
                    if precedent_count > 0 and precedent_count <= 5:
                        precedents = cell.get("directPrecedents", [])
                        precedents_str = _format_dependency_list(precedents, 3)
                        properties.append(f'prec=[{precedents_str}]')
                    elif precedent_count > 5:
                        properties.append(f'prec=[{precedent_count}refs]')
                    
                    if dependent_count > 0 and dependent_count <= 3:
                        dependents = cell.get("directDependents", [])
                        dependents_str = _format_dependency_list(dependents, 2)
                        properties.append(f'dept=[{dependents_str}]')
                    elif dependent_count > 3:
                        properties.append(f'dept=[{dependent_count}refs]')
            
            # Significant formatting (only if present)
            formatting = cell.get("formatting", {})
            if self._has_significant_formatting(formatting):
                format_parts = []
                
                # Font formatting
                font = formatting.get("font", {})
                if font.get("bold"):
                    format_parts.append("bold")
                if font.get("italic"):
                    format_parts.append("italic")
                if font.get("color") and font.get("color") not in [None, "auto", "#000000"]:
                    format_parts.append(f"color:{font.get('color')}")
                
                # Fill formatting
                fill = formatting.get("fill", {})
                if fill.get("startColor") and fill.get("startColor") not in [None, "auto", "#FFFFFF"]:
                    format_parts.append(f"fill:{fill.get('startColor')}")
                
                # Borders
                borders = formatting.get("borders", {})
                border_styles = []
                for side, border in borders.items():
                    if isinstance(border, dict) and border.get("style"):
                        border_styles.append(side[:1])  # Just first letter (l,r,t,b)
                if border_styles:
                    format_parts.append(f"border:{''.join(border_styles)}")
                
                # Number format
                num_format = formatting.get("numberFormat", "")
                if num_format and num_format.lower() not in ["general", "@"]:
                    format_parts.append(f"fmt:{num_format[:10]}")
                
                # Merged cells
                merged = formatting.get("merged", {})
                if merged.get("isMerged"):
                    format_parts.append("merged")
                
                if format_parts:
                    properties.append(f'fmt=[{",".join(format_parts)}]')
            
            # Data type (if not default)
            data_type = formatting.get("dataType", "") if formatting else ""
            if data_type and data_type not in ["n", ""]:
                properties.append(f'type={data_type}')
            
            # Comments (if present)
            if formatting and formatting.get("comment"):
                comment_text = formatting["comment"].get("text", "")[:20]
                if comment_text:
                    properties.append(f'comment={_safe_format_value(comment_text)}')
            
            # Hyperlinks (if present)
            if formatting and formatting.get("hyperlink"):
                hyperlink_target = formatting["hyperlink"].get("target", "")[:30]
                if hyperlink_target:
                    properties.append(f'link={_safe_format_value(hyperlink_target)}')
            
            return ", ".join(properties)
                
        except Exception as e:
            return f"ERROR: {str(e)[:30]}"

    def _is_significant_cell(self, cell: Dict[str, Any]) -> bool:
        """
        Determine if a cell has significant data worth including.
        Now includes cells with dependencies even if empty.
        
        Args:
            cell: Cell metadata dictionary
            
        Returns:
            True if cell should be included in compressed output
        """
        value = cell.get("value")
        formula = cell.get("formula")
        formatting = cell.get("formatting", {})
        
        # Include if has non-empty value
        if value is not None and str(value).strip() != "":
            return True
        
        # Include if has formula
        if formula and formula.startswith("="):
            return True
        
        # Include if has dependencies (new condition)
        if self.has_dependencies:
            total_connections = cell.get("totalConnections", 0)
            if total_connections > 0:
                return True
        
        # Include if has significant formatting (even if empty)
        if self._has_significant_formatting(formatting):
            return True
        
        return False

    def _has_significant_formatting(self, formatting: Dict[str, Any]) -> bool:
        """Check if formatting is significant enough to include - comprehensive version"""
        try:
            # Check font formatting
            font = formatting.get("font", {})
            if (font.get("bold") or font.get("italic") or font.get("color") or 
                font.get("underline") != "none" or font.get("strikethrough") or
                font.get("vertAlign") or (font.get("size") and font.get("size") != 11) or
                (font.get("name") and font.get("name") != "Calibri")):
                return True
            
            # Check fill formatting
            fill = formatting.get("fill", {})
            if (fill.get("startColor") or fill.get("endColor") or 
                fill.get("fillType") != "none" or fill.get("patternType")):
                return True
            
            # Check borders
            borders = formatting.get("borders", {})
            for border in borders.values():
                if isinstance(border, dict) and border.get("style"):
                    return True
            if borders.get("diagonalUp") or borders.get("diagonalDown") or borders.get("outline"):
                return True
            
            # Check alignment
            alignment = formatting.get("alignment", {})
            if (alignment.get("horizontal") or alignment.get("vertical") or
                alignment.get("wrapText") or alignment.get("shrinkToFit") or
                alignment.get("indent") != 0 or alignment.get("textRotation") != 0 or
                alignment.get("readingOrder") != 0 or alignment.get("justifyLastLine") or
                alignment.get("relativeIndent") != 0):
                return True
            
            # Check number format (if not general)
            number_format = formatting.get("numberFormat")
            if number_format and number_format.lower() not in ["general", "@"]:
                return True
            
            # Check protection
            protection = formatting.get("protection", {})
            if protection.get("locked") is False or protection.get("hidden"):
                return True
            
            # Check data type
            data_type = formatting.get("dataType", "")
            if data_type and data_type != "n":
                return True
            
            # Check merged cells
            merged = formatting.get("merged", {})
            if merged.get("isMerged"):
                return True
            
            # Check hyperlink
            if formatting.get("hyperlink"):
                return True
            
            # Check comment
            if formatting.get("comment"):
                return True
            
            return False
        except:
            return False

    def _escape_special_chars(self, text: str) -> str:
        """
        Escape special characters that could interfere with markdown parsing.
        
        Args:
            text: Text to escape
            
        Returns:
            Escaped text safe for markdown
        """
        if not isinstance(text, str):
            text = str(text)
        
        # Replace problematic characters
        replacements = {
            '#': '\\#',      # Markdown headers
            '*': '\\*',      # Markdown bold/italic
            '_': '\\_',      # Markdown italic/underline
            '`': '\\`',      # Markdown code
            '[': '\\[',      # Markdown links
            ']': '\\]',      # Markdown links
            '(': '\\(',      # Markdown links
            ')': '\\)',      # Markdown links
            '|': '\\|',      # Markdown tables
            '$': '\\$',      # LaTeX math
            '^': '\\^',      # Superscript
            '~': '\\~',      # Subscript
            '\\': '\\\\'     # Backslash
        }
        
        for char, replacement in replacements.items():
            text = text.replace(char, replacement)
        
        return text