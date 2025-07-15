import re
import json
from typing import List, Dict, Union, Optional, Any
from pathlib import Path
import sys
import os
import logging
import datetime
from decimal import Decimal
logger = logging.getLogger(__name__)
# Add the current directory to Python path
server_dir_path = Path(__file__).parent.parent.parent.parent.parent.absolute()
sys.path.append(str(server_dir_path))
from excel.editing.excel_writer import ExcelWriter

def update_excel_cache(workspace_path: str, all_updated_cells: List[Dict[str, Any]]) -> bool:
    """
    Update the Excel metadata cache with modified cell data.

    Args:
        workspace_path: Full path to the workbook in the format 'folder/workbook.xlsx'
        all_updated_cells: List of dicts with 'sheet_name' and 'updated_cells' keys.
                         Example: [{
                             "sheet_name": "Sheet1",
                             "updated_cells": [
                                 {"a": "A1", "f": "=SUM(B1:B2)", "v": 42},
                                 {"a": "B1", "f": "=5", "v": 5}
                             ]
                         }]

    Returns:
        bool: True if at least one cell was updated successfully, False otherwise
    """
    class ExtendedJSONEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, (datetime.datetime, datetime.date, datetime.time)):
                return obj.isoformat()
            elif isinstance(obj, Decimal):
                return str(obj)  # Preserve exact decimal precision
            return super().default(obj)

    MAPPINGS_FILE = server_dir_path / "metadata" / "__cache" / "files_mappings.json"
    
    try:
        with open(MAPPINGS_FILE, 'r') as f:
            mappings = json.load(f)
        temp_file_path = mappings.get(workspace_path) or next(
            (v for k, v in mappings.items() if k.endswith(Path(workspace_path).name)), 
            workspace_path
        )
    except (json.JSONDecodeError, OSError) as e:
        logger.error(f"Error loading file mappings: {str(e)}")
        temp_file_path = workspace_path

    try:
        cache_path = server_dir_path / "metadata" / "_cache" / "excel_metadata_hotcache.json"
        
        if not cache_path.exists():
            logger.error("Cache file not found")
            return False

        # Load the cache
        with open(cache_path, 'r+', encoding='utf-8') as f:
            try:
                cache_data = json.load(f)
            except json.JSONDecodeError as e:
                logger.error(f"Invalid cache file format: {str(e)}")
                return False

            file_name = os.path.basename(temp_file_path)
            workbook_updated = False
            success_count = 0
            error_count = 0
            
            for cache_key, workbook_data in cache_data.items():
                if not isinstance(workbook_data, dict) or workbook_data.get('workbook_name') != file_name:
                    continue
                    
                workbook_updated = True
                
                # Process each sheet's updates
                for sheet_update in all_updated_cells:
                    sheet_name = sheet_update.get('sheet_name')
                    updated_cells = sheet_update.get('updated_cells', [])
                    
                    if not sheet_name or not updated_cells:
                        continue
                    
                    # Initialize sheet data if it doesn't exist
                    if "sheets" not in workbook_data:
                        workbook_data["sheets"] = {}
                        
                    if sheet_name not in workbook_data["sheets"]:
                        workbook_data["sheets"][sheet_name] = {"chunks": [{"cells": []}]}
                    
                    sheet_data = workbook_data["sheets"][sheet_name]
                    
                    # Ensure chunks exist
                    if "chunks" not in sheet_data or not sheet_data["chunks"]:
                        sheet_data["chunks"] = [{"cells": []}]
                    
                    # Ensure first chunk has cells list
                    if "cells" not in sheet_data["chunks"][0]:
                        sheet_data["chunks"][0]["cells"] = []
                    
                    # Create mapping of cell references to their indices
                    existing_cells = {
                        cell.get('a'): idx 
                        for idx, cell in enumerate(sheet_data["chunks"][0]["cells"])
                    }
                    
                    # Process each cell update
                    for cell_data in updated_cells:
                        try:
                            cell_ref = cell_data.get('a')
                            if not cell_ref:
                                error_count += 1
                                continue
                            
                            cell_entry = {
                                'a': cell_ref,
                                'f': cell_data.get('f'),
                                'v': cell_data.get('v')
                            }
                            
                            # Update existing cell or add new one
                            if cell_ref in existing_cells:
                                sheet_data["chunks"][0]["cells"][existing_cells[cell_ref]] = cell_entry
                                #logger.info(f"Updated existing cell: {cell_ref}")
                            else:
                                sheet_data["chunks"][0]["cells"].append(cell_entry)
                                #logger.info(f"Added new cell: {cell_ref}")
                            
                            success_count += 1
                            
                        except Exception as cell_error:
                            error_count += 1
                            logger.error(f"Error updating cell {cell_data.get('a', 'unknown')}: {str(cell_error)}", 
                                       exc_info=True)
                            continue
            
            if not workbook_updated:
                logger.error(f"No matching workbook found for {file_name}")
                return False
                
            if success_count == 0 and error_count > 0:
                logger.error("All cell updates failed")
                return False
                
            # Write back to the file
            try:
                f.seek(0)
                json.dump(cache_data, f, indent=2, cls=ExtendedJSONEncoder)
                f.truncate()
                logger.info(f"Cache updated successfully: {success_count} cells updated, {error_count} errors")
                return True
            except Exception as write_error:
                logger.error(f"Error writing to cache file: {str(write_error)}", exc_info=True)
                return False
                
    except Exception as e:
        logger.error(f"Error updating cache: {str(e)}", exc_info=True)
        return False

def get_simplified_excel_metadata(workspace_path: str) -> str:
    """
    Retrieve simplified metadata for the specified excel file from the hotcache.
    Only returns address, value, and formula for each cell.
    
    Args:
        workspace_path: Full path to the excel file workbook in the format 'folder/workbook.xlsx'
    
    Returns:
        A JSON string containing simplified metadata with only address, value, and formula:
        {
            "Sheet1": [
                {
                    "a": "A1", 
                    "f": "=SUM(B1:B2)", 
                    "v": 42
                }
            ],
            "Sheet2": [...]
        }
    """
    
    MAPPINGS_FILE = server_dir_path / "metadata" / "__cache" / "files_mappings.json"
    
    try:
        with open(MAPPINGS_FILE, 'r') as f:
            mappings = json.load(f)
        temp_file_path = mappings.get(workspace_path) or next(
            (v for k, v in mappings.items() if k.endswith(Path(workspace_path).name)), 
            workspace_path
        )
    except (json.JSONDecodeError, OSError):
        temp_file_path = workspace_path

    try:
        # Get the cache file
        cache_path = server_dir_path / "metadata" / "_cache" / "excel_metadata_hotcache.json"
        
        if not cache_path.exists():
            return 'Cache file not found'

        # Load the cache
        with open(cache_path, 'r') as f:
            cache_data = json.load(f)

        # Find the workbook by matching the workbook_name
        file_name = os.path.basename(temp_file_path)
        workbook_data = None
        
        for cache_key, data in cache_data.items():
            if isinstance(data, dict) and data.get('workbook_name') == file_name:
                workbook_data = data
                break
        
        if not workbook_data:
            return 'Workbook not found in cache'
            
        # Extract simplified data
        simplified_data = {}
        
        if 'sheets' in workbook_data:
            for sheet_name, sheet_data in workbook_data['sheets'].items():
                simplified_data[sheet_name] = []
                
                # Process chunks if they exist
                if 'chunks' in sheet_data:
                    for chunk in sheet_data['chunks']:
                        if 'cells' in chunk:
                            for cell in chunk['cells']:
                                # Only include address, value, and formula
                                simplified_cell = {}
                                
                                # Address (always include)
                                if 'a' in cell:
                                    simplified_cell['a'] = cell['a']
                                
                                # Value (only if not None)
                                if 'v' in cell and cell['v'] is not None:
                                    simplified_cell['v'] = cell['v']
                                
                                # Formula (only if not None)
                                if 'f' in cell and cell['f'] is not None:
                                    simplified_cell['f'] = cell['f']
                                
                                # Only add cell if it has at least an address
                                if 'a' in simplified_cell:
                                    simplified_data[sheet_name].append(simplified_cell)
        
        return json.dumps(simplified_data, separators=(',', ':'))
        
    except json.JSONDecodeError:
        return 'Failed to parse cache file'
    except Exception as e:
        logger.error(f"Error retrieving simplified data: {str(e)}", exc_info=True)
        return 'Failed to get simplified data from cache'


def get_full_excel_metadata(workspace_path: str) -> str:
    """
    Retrieve complete metadata for the specified excel file from the hotcache.
    
    Args:
        workspace_path: Full path to the excel file workbook in the format 'folder/workbook.xlsx'
    
    Returns:
        A JSON string containing all metadata for the excel file
    """

    MAPPINGS_FILE = server_dir_path / "metadata" / "__cache" / "files_mappings.json"
    
    try:
        with open(MAPPINGS_FILE, 'r') as f:
            mappings = json.load(f)
        temp_file_path = mappings.get(workspace_path) or next(
            (v for k, v in mappings.items() if k.endswith(Path(workspace_path).name)), 
            workspace_path
        )
    except (json.JSONDecodeError, OSError):
        temp_file_path = workspace_path

    try:
        # Get the cache file
        cache_path = server_dir_path / "metadata" / "_cache" / "excel_metadata_hotcache.json"
        
        if not cache_path.exists():
            return 'Cache file not found'

        # Load the cache
        with open(cache_path, 'r') as f:
            cache_data = json.load(f)


        # Find the workbook by matching the workbook_name
        file_name = os.path.basename(temp_file_path)
        for cache_key, workbook_data in cache_data.items():
            if isinstance(workbook_data, dict) and workbook_data.get('workbook_name') == file_name:
                return workbook_data
        

        # Return the complete workbook data
        return json.dumps(workbook_data, separators=(',', ':'))
        
    except json.JSONDecodeError:
        return 'Failed to parse cache file'
    except Exception as e:
        logger.error(f"Error retrieving data: {str(e)}", exc_info=True)
        return 'Failed to get data from cache'

def get_full_metadata_from_cache(workspace_path: str, sheet_cell_ranges: Optional[Dict[str, List[str]]] = None) -> str:
    """
    Retrieve complete metadata including formatting for the specified excel file from the hotcache,
    returning only the specified cell ranges with all non-null/non-false formatting properties.
    
    Args:
        workspace_path: Full path to the excel file workbook in the format 'folder/workbook.xlsx'
        sheet_cell_ranges: Dict mapping sheet names to lists of cell ranges.
                         If None, returns all data with complete formatting.
                         Example: {"Sheet1": ["A1:B10", "C1:D5"], "Sheet2": ["A1:Z1000"]}
    
    Returns:
        A JSON string containing complete metadata with formatting in the format:
        {
            "Sheet1": [
                {
                    "a": "A1", 
                    "f": "=SUM(B1:B2)", 
                    "v": 42,
                    "fmt": {
                        "font": {"name": "Arial", "size": 12, "bold": true},
                        "numberFormat": "General"
                    }
                }
            ],
            "Sheet2": [...]
        }
    """
    def normalize_cell_ref(cell_ref: str) -> str:
        """Remove $ signs from cell reference and convert to uppercase."""
        if not cell_ref:
            return cell_ref
        return cell_ref.replace('$', '').upper()
    
    def filter_formatting(fmt_data: dict) -> dict:
        """Filter out null/false values and skip protection/dataType properties."""
        if not isinstance(fmt_data, dict):
            return {}
        
        filtered = {}
        
        # Skip protection and dataType properties entirely
        skip_properties = {'protection', 'dataType'}
        
        for key, value in fmt_data.items():
            if key in skip_properties:
                continue
                
            if isinstance(value, dict):
                # Recursively filter nested dictionaries
                filtered_nested = filter_formatting(value)
                if filtered_nested:  # Only add if there are non-null values
                    filtered[key] = filtered_nested
            elif value is not None and value is not False and value != "" and value != 0:
                # Include non-null, non-false, non-empty values (but allow 0.0 for float values)
                if isinstance(value, (int, float)) and value == 0:
                    # Skip zero values for formatting properties
                    continue
                filtered[key] = value
                
        return filtered
    
    def process_cell_data(cell: dict) -> dict:
        """Process a single cell and return it with filtered formatting."""
        result = {'a': cell.get('a')}  # Always include address
        
        # Add basic properties if they exist and are not null
        for prop in ['v', 'f', 'r', 'c']:
            if prop in cell and cell[prop] is not None:
                result[prop] = cell[prop]
        
        # Process formatting if it exists
        if 'fmt' in cell and isinstance(cell['fmt'], dict):
            filtered_fmt = filter_formatting(cell['fmt'])
            if filtered_fmt:  # Only add if there are meaningful formatting properties
                result['fmt'] = filtered_fmt
        
        # Add dependency data if present
        for prop in ['precedents', 'dependents', 'precedentCount', 'dependentCount', 'totalConnections']:
            if prop in cell and cell[prop] is not None:
                if isinstance(cell[prop], (list, int)) and cell[prop]:  # Non-empty lists or non-zero numbers
                    result[prop] = cell[prop]
        
        return result

    try:
        # First get the full metadata from cache
        full_metadata = get_full_excel_metadata(workspace_path)
        if not full_metadata or not isinstance(full_metadata, dict):
            return json.dumps({"error": "Failed to load metadata from cache"})
        
        # If no ranges specified, return all data with complete formatting
        if not sheet_cell_ranges:
            result = {}
            for sheet_name, sheet_data in full_metadata.get('sheets', {}).items():
                result[sheet_name] = []
                for chunk in sheet_data.get('chunks', []):
                    for cell in chunk.get('cells', []):
                        processed_cell = process_cell_data(cell)
                        result[sheet_name].append(processed_cell)
            return json.dumps(result, separators=(',', ':'))
        
        # Parse cell range strings into cell references
        def parse_cell_refs(cell_range: str) -> List[str]:
            """Parse a cell range into individual cell references."""
            import re
            from openpyxl.utils import range_boundaries, get_column_letter
            
            # Handle single cell
            if ':' not in cell_range:
                return [normalize_cell_ref(cell_range)]
                
            # Handle range
            try:
                # Remove $ signs before parsing range
                normalized_range = cell_range.replace('$', '')
                min_col, min_row, max_col, max_row = range_boundaries(normalized_range)
                cells = []
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell_ref = f"{get_column_letter(col)}{row}"
                        cells.append(cell_ref)
                return cells
            except Exception as e:
                logger.warning(f"Invalid range format '{cell_range}': {str(e)}")
                return []
        
        # Filter the data based on the specified ranges
        filtered_data = {}
        for sheet_name, ranges in sheet_cell_ranges.items():
            if sheet_name not in full_metadata.get('sheets', {}):
                continue

            if isinstance(ranges, str):
                ranges = [ranges]
                
            # Get all cell references to include
            cell_refs_to_include = set()
            for r in ranges:
                cell_refs_to_include.update(parse_cell_refs(r))
                
            if not cell_refs_to_include:
                continue
                
            # Filter cells
            sheet_data = full_metadata['sheets'][sheet_name]
            filtered_cells = []
            cell_refs_found = set()
            
            for chunk in sheet_data.get('chunks', []):
                for cell in chunk.get('cells', []):
                    cell_ref = cell.get('a')
                    normalized_ref = normalize_cell_ref(cell_ref)
                    if cell_ref and normalized_ref in cell_refs_to_include:
                        processed_cell = process_cell_data(cell)
                        filtered_cells.append(processed_cell)
                        cell_refs_found.add(normalized_ref)
            
            # Log any requested but not found cells
            missing_cells = cell_refs_to_include - cell_refs_found
            if missing_cells:
                logger.debug(f"{len(missing_cells)}Cells not found in sheet '{sheet_name}'")
                
            if filtered_cells:
                filtered_data[sheet_name] = filtered_cells
        
        return json.dumps(filtered_data, separators=(',', ':'))

    except Exception as e:
        logger.error(f"Error getting full metadata with formatting: {str(e)}", exc_info=True)
        return json.dumps({"error": f"Failed to get full metadata: {str(e)}"})

def get_metadata_from_cache(workspace_path: str, sheet_cell_ranges: Optional[Dict[str, List[str]]] = None) -> str:
    """
    Retrieve filtered metadata for the specified excel file from the hotcache,
    returning only the specified cell ranges without row range validation.
    
    Args:
        workspace_path: Full path to the excel file workbook in the format 'folder/workbook.xlsx'
        sheet_cell_ranges: Dict mapping sheet names to lists of cell ranges.
                         If None, returns all data (same as get_full_excel_metadata).
                         Example: {"Sheet1": ["A1:B10", "C1:D5"], "Sheet2": ["A1:Z1000"]}
    
    Returns:
        A JSON string containing filtered metadata in the format:
        {
            "Sheet1": [
                {"a": "A1", "f": "=SUM(B1:B2)", "v": 42},
                {"a": "B1", "f": "=5", "v": 5}
            ],
            "Sheet2": [...]
        }
    """
    def normalize_cell_ref(cell_ref: str) -> str:
        """Remove $ signs from cell reference and convert to uppercase."""
        if not cell_ref:
            return cell_ref
        return cell_ref.replace('$', '').upper()

    try:
        # First get the full metadata from cache
        full_metadata = get_full_excel_metadata(workspace_path)
        if not full_metadata or not isinstance(full_metadata, dict):
            return json.dumps({"error": "Failed to load metadata from cache"})
        
        # If no ranges specified, return all data in the expected format
        if not sheet_cell_ranges:
            result = {}
            for sheet_name, sheet_data in full_metadata.get('sheets', {}).items():
                result[sheet_name] = []
                for chunk in sheet_data.get('chunks', []):
                    result[sheet_name].extend(chunk.get('cells', []))
            return json.dumps(result, separators=(',', ':'))
        
        # Parse cell range strings into cell references
        def parse_cell_refs(cell_range: str) -> List[str]:
            """Parse a cell range into individual cell references."""
            import re
            from openpyxl.utils import range_boundaries, get_column_letter
            
            # Handle single cell
            if ':' not in cell_range:
                return [normalize_cell_ref(cell_range)]
                
            # Handle range
            try:
                # Remove $ signs before parsing range
                normalized_range = cell_range.replace('$', '')
                min_col, min_row, max_col, max_row = range_boundaries(normalized_range)
                cells = []
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell_ref = f"{get_column_letter(col)}{row}"
                        cells.append(cell_ref)
                return cells
            except Exception as e:
                logger.warning(f"Invalid range format '{cell_range}': {str(e)}")
                return []
        
        # Filter the data based on the specified ranges
        filtered_data = {}
        for sheet_name, ranges in sheet_cell_ranges.items():
            if sheet_name not in full_metadata.get('sheets', {}):
                continue

            if isinstance(ranges, str):
                ranges = [ranges]
                
            # Get all cell references to include
            cell_refs_to_include = set()
            for r in ranges:
                cell_refs_to_include.update(parse_cell_refs(r))
                
            if not cell_refs_to_include:
                continue
                
            # Filter cells
            sheet_data = full_metadata['sheets'][sheet_name]
            filtered_cells = []
            cell_refs_found = set()
            
            for chunk in sheet_data.get('chunks', []):
                for cell in chunk.get('cells', []):
                    cell_ref = cell.get('a')
                    normalized_ref = normalize_cell_ref(cell_ref)
                    if cell_ref and normalized_ref in cell_refs_to_include:
                        filtered_cells.append({
                            'a': cell_ref,  # Keep original reference with $ if present
                            'f': cell.get('f'),
                            'v': cell.get('v')
                        })
                        cell_refs_found.add(normalized_ref)
            
            # Log any requested but not found cells
            missing_cells = cell_refs_to_include - cell_refs_found
            if missing_cells:
                logger.debug(f"{len(missing_cells)}Cells not found in sheet '{sheet_name}'")
                
            if filtered_cells:
                filtered_data[sheet_name] = filtered_cells
        
        return json.dumps(filtered_data, separators=(',', ':'))

    except Exception as e:
        logger.error(f"Error getting filtered metadata: {str(e)}", exc_info=True)
        return json.dumps({"error": f"Failed to get filtered metadata: {str(e)}"})

def get_excel_metadata(workspace_path: str, sheet_cell_ranges: Optional[Dict[str, List[str]]] = None) -> str:
    """
    Retrieve complete metadata for the specified excel file using xlwings.
    
    Args:
        workspace_path: Full path to the excel file
        sheet_cell_ranges: Optional dict mapping sheet names to lists of cell ranges.
                            If None, returns all used cells in all sheets.
                            Example: {"Sheet1": ["A1:B10", "C1:D5"], "Sheet2": ["A1:Z1000"]}
        
    Returns:
        A JSON string containing sheet names as keys and lists of cell data as values
        Format: {sheet_name: [{"a": "A1", "f": "=SUM(A2:A5)", "v": "100"}, ...], ...}
    """

    MAPPINGS_FILE = server_dir_path / "metadata" / "__cache" / "files_mappings.json"
    
    try:
        with open(MAPPINGS_FILE, 'r') as f:
            mappings = json.load(f)
        temp_file_path = mappings.get(workspace_path) or next(
            (v for k, v in mappings.items() if k.endswith(Path(workspace_path).name)), 
            workspace_path
        )
    except (json.JSONDecodeError, OSError):
        temp_file_path = workspace_path

    try:
        with ExcelWriter(visible=True) as writer:
            # Get all data from all sheets
            result = writer.get_workbook_metadata(temp_file_path, sheet_cell_ranges)
            
            # Transform the data to match the expected format
        if result:
            formatted_data = {}
            for sheet_name, cells in result.get("data", {}).items():
                formatted_cells = []
                for cell in cells:
                    cell_data = {
                        "a": cell["address"],
                        "v": str(cell["value"]) if cell["value"] is not None else ""
                    }
                    if cell.get("formula"):
                        cell_data["f"] = cell["formula"]
                    formatted_cells.append(cell_data)
                
                if formatted_cells:
                    formatted_data[sheet_name] = formatted_cells
            
            # Add error information if any
            for sheet_name, errors in result.get("errors", {}).items():
                if sheet_name not in formatted_data:
                    formatted_data[sheet_name] = []
                
                for error in errors:
                    formatted_data[sheet_name].append({
                        "a": error["cell"],
                        "v": f"#ERROR: {error['error']}",
                        "f": error["formula"]
                    })
            
            return json.dumps(formatted_data, separators=(',', ':'))
        else:
            return json.dumps({'error': 'Failed to get cell data. Must do without.'})    
            
        
    except Exception as e:
        logger.error(f"Error getting cell data: {str(e)}", exc_info=True)
        return [{"error": f"Failed to get cell data: {str(e)}"}]




def get_cell_formulas_from_cache(workspace_path: str, cell_dict: Dict[str, Dict[str, str]]) -> Optional[Dict[str, Dict[str, str]]]:
    """
    Get formulas from specific cells in an Excel workbook from the cache.
    
    Args:
        workspace_path: Path to the Excel file
        cell_dict: Dictionary mapping sheet names to cell references
            Example: {
                "Sheet1": {"A1": "=SUM(B1:B10)", "B1": "=A1*2"},
                "Sheet2": {"C1": "=AVERAGE(A1:A10)"}
            }
            Note: The formula values are ignored, only cell references are used.
            
    Returns:
        Dictionary with the same structure as input, containing the actual formulas from the Excel cache.
        Returns None if there was an error accessing the cache.
        For cells not found in cache, returns empty string as the formula.
    """
    def normalize_cell_ref(cell_ref: str) -> str:
        """Remove $ signs from cell reference and convert to uppercase."""
        if not cell_ref:
            return cell_ref
        return cell_ref.replace('$', '').upper()

    try:
        # Get the full metadata from cache
        full_metadata = get_full_excel_metadata(workspace_path)
        if not full_metadata or not isinstance(full_metadata, dict):
            logger.error("Failed to load metadata from cache")
            return None

        result = {}
        
        # Process each sheet in the request
        for sheet_name, cells in cell_dict.items():
            result[sheet_name] = {}
            
            # Initialize all requested cells with empty strings
            for cell_ref in cells:
                result[sheet_name][cell_ref] = ""
                
            # If sheet exists in cache, try to find the cells
            if sheet_name in full_metadata.get('sheets', {}):
                sheet_data = full_metadata['sheets'][sheet_name]
                
                # Create a mapping of normalized cell refs to original cell refs
                cell_refs = {normalize_cell_ref(ref): ref for ref in cells.keys()}
                found_cells = set()
                
                # Search through all chunks for the requested cells
                for chunk in sheet_data.get('chunks', []):
                    for cell in chunk.get('cells', []):
                        cell_ref = cell.get('a')
                        if not cell_ref:
                            continue
                            
                        normalized_ref = normalize_cell_ref(cell_ref)
                        if normalized_ref in cell_refs:
                            # Use the original casing from the request for the result
                            original_ref = cell_refs[normalized_ref]
                            result[sheet_name][original_ref] = cell.get('f', '')
                            found_cells.add(normalized_ref)
                
                # Log any missing cells
                missing_cells = set(cell_refs.keys()) - found_cells
                if missing_cells:
                    logger.debug(f"{len(missing_cells)}Cells not found in sheet '{sheet_name}'")
            else:
                logger.warning(f"Sheet '{sheet_name}' not found in cache")
                # Keep the empty strings for all cells in this sheet
                
        return result

    except Exception as e:
        logger.error(f"Error getting cell formulas from cache: {str(e)}", exc_info=True)
        return None

def get_cell_formulas(workspace_path: str, cell_dict: Dict[str, Dict[str, str]] = None) -> str:
    """
            Get formulas from specific cells in an Excel workbook.
            
            Args:
                file_path: Path to the Excel file
                cell_dict: Dictionary mapping sheet names to cell references
                    Example: {
                        "Sheet1": {"A1": "=SUM(B1:B10)", "B1": "=A1*2"},
                        "Sheet2": {"C1": "=AVERAGE(A1:A10)"}
                    }
                    Note: The formula values are ignored, only cell references are used.
                    
            Returns:
                Dictionary with the same structure as input, but containing the actual
                formulas from the Excel file.
                Example: {
                    "Sheet1": {"A1": "=SUM(B1:B10)", "B1": "=A1*2"},
                    "Sheet2": {"C1": "=AVERAGE(A1:A10)"}
                }
            """

    MAPPINGS_FILE = server_dir_path / "metadata" / "__cache" / "files_mappings.json"
    
    try:
        with open(MAPPINGS_FILE, 'r') as f:
            mappings = json.load(f)
        temp_file_path = mappings.get(workspace_path) or next(
            (v for k, v in mappings.items() if k.endswith(Path(workspace_path).name)), 
            workspace_path
        )
    except (json.JSONDecodeError, OSError):
        temp_file_path = workspace_path

    try:
        with ExcelWriter(visible=True) as writer:
            result = writer.get_cell_formulas(temp_file_path, cell_dict)
        if result:
            return result
        else:
            return {} 
            
        
    except Exception as e:
        logger.error(f"Error getting cell data: {str(e)}", exc_info=True)
        return [{"error": f"Failed to get cell data: {str(e)}"}]


def xl_col_to_name(col_num):
    """Convert a column number to Excel column name (A, B, ..., Z, AA, AB, ...)"""
    col_name = ''
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        col_name = chr(65 + remainder) + col_name
    return col_name


def get_formulas_for_revert(
    sheet_formulas: Dict[str, Dict[str, str]], 
    sheet_cell_ranges: Dict[str, List[str]]
) -> Dict[str, Dict[str, str]]:
    """
    Extract formulas from a dictionary of sheet data based on specified cell ranges.
    
    Args:
        sheet_formulas: Dictionary mapping sheet names to cell formulas.
                      Example: {
                          "Sheet1": {"A1": "=SUM(B1:B10)", "B1": "=A1*2"},
                          "Sheet2": {"C1": "=AVERAGE(A1:A10)"}
                      }
        sheet_cell_ranges: Dict mapping sheet names to lists of cell ranges.
                         Example: {"Sheet1": ["A1:B10", "C1:D5"], "Sheet2": ["A1:Z1000"]}
    
    Returns:
        A dictionary with the same structure as sheet_formulas, containing only the
        cells that fall within the specified ranges.
        Example: {
            "Sheet1": {"A1": "=SUM(B1:B10)", "B1": "=A1*2"},
            "Sheet2": {"C1": "=AVERAGE(A1:A10)"}
        }
    """
    def normalize_cell_ref(cell_ref: str) -> str:
        """Remove $ signs from cell reference and convert to uppercase."""
        if not cell_ref:
            return cell_ref
        return cell_ref.replace('$', '').upper()

    def parse_cell_refs(cell_range: str) -> List[str]:
        """Parse a cell range into individual cell references."""
        import re
        from openpyxl.utils import range_boundaries, get_column_letter
        
        # Handle single cell
        if ':' not in cell_range:
            return [normalize_cell_ref(cell_range)]
            
        # Handle range
        try:
            # Remove $ signs before parsing range
            normalized_range = cell_range.replace('$', '')
            min_col, min_row, max_col, max_row = range_boundaries(normalized_range)
            cells = []
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell_ref = f"{get_column_letter(col)}{row}"
                    cells.append(cell_ref)
            return cells
        except Exception as e:
            logger.warning(f"Invalid range format '{cell_range}': {str(e)}")
            return []

    try:
        result = {}
        
        for sheet_name, ranges in sheet_cell_ranges.items():
            if sheet_name not in sheet_formulas:
                logger.debug(f"Sheet '{sheet_name}' not found in provided formulas")
                continue

            if isinstance(ranges, str):
                ranges = [ranges]
                
            # Get all cell references to include
            cell_refs_to_include = set()
            for r in ranges:
                cell_refs_to_include.update(parse_cell_refs(r))
                
            if not cell_refs_to_include:
                logger.debug(f"No valid cell ranges found for sheet '{sheet_name}'")
                continue
                
            # Initialize sheet in result
            if sheet_name not in result:
                result[sheet_name] = {}
                
            # Filter cells
            sheet_data = sheet_formulas[sheet_name]
            cell_refs_found = set()
            
            for cell_ref, formula in sheet_data.items():
                normalized_ref = normalize_cell_ref(cell_ref)
                if normalized_ref in cell_refs_to_include:
                    result[sheet_name][cell_ref] = formula  # Keep original cell reference
                    cell_refs_found.add(normalized_ref)
            
            # Log any requested but not found cells
            missing_cells = cell_refs_to_include - cell_refs_found
            if missing_cells:
                logger.debug(f"{len(missing_cells)}Cells not found in sheet '{sheet_name}'")
        
        return result

    except Exception as e:
        logger.error(f"Error extracting formulas: {str(e)}", exc_info=True)
        return {"error": f"Failed to extract formulas: {str(e)}"}


def clean_json_string(json_str):
    """Clean and parse a JSON string that might be malformed or have extra escaping.
    
    Handles various edge cases including:
    - Extra outer quotes
    - Escaped quotes
    - Truncated JSON
    - Malformed structures
    - Single-quoted strings
    - Trailing commas
    
    Args:
        json_str: A string that might be a JSON string, possibly with extra escaping
                or malformed structure. Expected format: {sheet: {cell1:val, ...}, ...}
        
    Returns:
        Parsed Python object from the JSON, or the original string if all parsing fails
    """
    import re
    import json
    import ast
    from json import JSONDecodeError

    if not isinstance(json_str, str) or not json_str.strip():
        return json_str

    # Helper function to safely parse JSON
    def try_parse(s):
        try:
            return json.loads(s)
        except (JSONDecodeError, TypeError):
            return None

    # Try direct JSON parse first (handles well-formed JSON)
    parsed = try_parse(json_str)
    if parsed is not None:
        return parsed

    # Handle case where the entire JSON is wrapped in quotes and escaped
    stripped = json_str.strip()
    if (stripped.startswith('"') and stripped.endswith('"')) or \
       (stripped.startswith("'") and stripped.endswith("'")):
        try:
            # Remove outer quotes and unescape
            unescaped = stripped[1:-1].encode().decode('unicode_escape')
            # Try parsing the unescaped content
            parsed = try_parse(unescaped)
            if parsed is not None:
                return parsed
            # If unescaped content is still not valid JSON, try wrapping in braces
            if not unescaped.startswith('{'):
                parsed = try_parse(f'{{{unescaped}}}')
                if parsed is not None:
                    return parsed
        except (UnicodeDecodeError, JSONDecodeError):
            pass

    # Try to extract valid JSON using regex for the expected structure
    try:
        # Pattern for sheet with cell ranges: {"Sheet1": ["A1:B5"]} or similar
        sheet_range_pattern = r'\{[^{}]*"[^"]*"\s*:\s*(\[\s*"[^"]*"\s*(?:,\s*"[^"]*"\s*)*\]|\{[^{}]*\})[^{}]*\}'
        matches = re.findall(sheet_range_pattern, json_str, re.DOTALL)
        if matches:
            # Take the longest match that looks like valid JSON
            best_match = max(matches, key=len)
            # Try to balance the braces if needed
            open_braces = best_match.count('{')
            close_braces = best_match.count('}')
            if open_braces > close_braces:
                best_match += '}' * (open_braces - close_braces)
            elif close_braces > open_braces:
                best_match = '{' * (close_braces - open_braces) + best_match
            parsed = try_parse(best_match)
            if parsed is not None:
                return parsed
    except (re.error, JSONDecodeError):
        pass

    # Try to fix common JSON issues
    try:
        # Remove control characters
        cleaned = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', json_str)
        # Convert single quotes to double quotes
        cleaned = re.sub(r"(?<!\\)'", '"', cleaned)
        # Handle escaped single quotes
        cleaned = re.sub(r"(?<!\\)\\'", "'", cleaned)
        # Remove trailing commas
        cleaned = re.sub(r',\s*([}\]])', r'\1', cleaned)
        # Remove comments (// and /* */)
        cleaned = re.sub(r'//.*?$|/\*.*?\*/', '', cleaned, flags=re.MULTILINE)
        
        parsed = try_parse(cleaned)
        if parsed is not None:
            return parsed
    except (re.error, JSONDecodeError):
        pass

    # Try to extract cell data as last resort
    try:
        # Look for patterns like "A1": "value" or "A1":"value"
        cell_pattern = r'"([A-Z]+\d+)"\s*:\s*("[^"]*"|[\d.]+|true|false|null)'
        sheet_pattern = r'"([^"]+)"\s*:\s*(\{[^{}]*\}|\[[^]]*\])'
        
        # Find all sheet matches
        sheet_matches = re.findall(sheet_pattern, json_str, re.DOTALL)
        if sheet_matches:
            result = {}
            for sheet_name, cells_str in sheet_matches:
                # If it's a list of ranges like ["A1:B5", "C1:D5"]
                if cells_str.startswith('['):
                    ranges = re.findall(r'"([^"]+)"', cells_str)
                    if ranges:
                        result[sheet_name] = ranges
                # If it's an object with cell values
                else:
                    cell_matches = re.findall(cell_pattern, cells_str)
                    if cell_matches:
                        result[sheet_name] = dict(cell_matches)
            if result:
                return result
    except (re.error, TypeError):
        pass

    # If all else fails, return the original string
    return json_str